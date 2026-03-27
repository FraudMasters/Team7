"""
Report Template Service for generating various report types.

This module provides functionality to generate different types of reports
(candidate summary, hiring pipeline, EEOC compliance, custom analytics)
in multiple formats (PDF, Excel, CSV).

The service supports:
- Candidate summary reports with resume data, match scores, and hiring stage
- Hiring pipeline reports with stage breakdowns, time metrics, and conversion rates
- EEOC compliance reports with diversity metrics and demographic breakdowns
- Custom analytics reports with user-selected metrics and filters
- Multiple export formats: PDF (via reportlab), Excel (via openpyxl), CSV

Report types:
- candidate_summary: Individual candidate report with detailed information
- hiring_pipeline: Pipeline analysis with stage metrics and conversion rates
- eeoc_compliance: EEOC compliance report with diversity metrics
- custom_analytics: Custom report with user-selected metrics

The service integrates with:
- ReportGenerator for PDF generation
- Export services for Excel/CSV generation
- Database models for data collection
"""
import io
import logging
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Union
from uuid import UUID

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from config import get_settings

logger = logging.getLogger(__name__)

# Global report template service instance
_report_template_service: Optional["ReportTemplateService"] = None


@dataclass
class ReportConfig:
    """
    Configuration for report generation.

    Attributes:
        report_type: Type of report (candidate_summary, hiring_pipeline, eeoc_compliance, custom_analytics)
        output_format: Output format (pdf, excel, csv)
        title: Report title
        organization_id: Organization ID for filtering data
        filters: Additional filters for data collection
        metrics: List of metrics to include (for custom reports)
        include_charts: Whether to include charts/visualizations
        page_format: Page format for PDF (A4, Letter)
        date_range: Optional date range for filtering
    """
    report_type: str
    output_format: str = "pdf"
    title: Optional[str] = None
    organization_id: Optional[str] = None
    filters: Dict[str, Any] = field(default_factory=dict)
    metrics: List[str] = field(default_factory=list)
    include_charts: bool = True
    page_format: str = "A4"
    date_range: Optional[Dict[str, str]] = None

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for JSON serialization."""
        return {
            "report_type": self.report_type,
            "output_format": self.output_format,
            "title": self.title,
            "organization_id": self.organization_id,
            "filters": self.filters,
            "metrics": self.metrics,
            "include_charts": self.include_charts,
            "page_format": self.page_format,
            "date_range": self.date_range,
        }


@dataclass
class ReportGenerationResult:
    """
    Result of report generation.

    Attributes:
        success: Whether report generation succeeded
        report_bytes: Generated report content as bytes
        filename: Suggested filename for the report
        content_type: MIME type (application/pdf, application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, text/csv)
        error_message: Error message if generation failed
        file_size: Size of generated report in bytes
        metadata: Additional metadata about the report
        report_type: Type of report that was generated
        generated_at: Timestamp when report was generated
    """
    success: bool
    report_bytes: Optional[bytes] = None
    filename: Optional[str] = None
    content_type: str = "application/pdf"
    error_message: Optional[str] = None
    file_size: int = 0
    metadata: Dict[str, Any] = field(default_factory=dict)
    report_type: Optional[str] = None
    generated_at: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for JSON serialization."""
        return {
            "success": self.success,
            "filename": self.filename,
            "content_type": self.content_type,
            "error_message": self.error_message,
            "file_size": self.file_size,
            "metadata": self.metadata,
            "report_type": self.report_type,
            "generated_at": self.generated_at,
        }


class ReportTemplateService:
    """
    Report Template Service for generating various report types.

    This service generates professional reports in multiple formats:
    - PDF reports using reportlab
    - Excel reports using openpyxl
    - CSV exports for data portability

    The service supports multiple report types:
    - candidate_summary: Individual candidate reports with detailed information
    - hiring_pipeline: Pipeline analysis with stage metrics and conversion rates
    - eeoc_compliance: EEOC compliance reports with diversity metrics
    - custom_analytics: Custom reports with user-selected metrics

    Attributes:
        db: Database session for data collection
        enabled: Whether report generation is enabled
        default_page_format: Default page format for PDF reports

    Example:
        >>> service = ReportTemplateService(db)
        >>> config = ReportConfig(
        ...     report_type="candidate_summary",
        ...     output_format="pdf",
        ...     title="Candidate Summary Report"
        ... )
        >>> result = await service.generate_report(
        ...     config=config,
        ...     data={"resume_id": "abc-123", ...}
        ... )
        >>> if result.success:
        ...     with open("report.pdf", "wb") as f:
        ...         f.write(result.report_bytes)
    """

    # Report type constants
    REPORT_TYPE_CANDIDATE_SUMMARY = "candidate_summary"
    REPORT_TYPE_HIRING_PIPELINE = "hiring_pipeline"
    REPORT_TYPE_EEOC_COMPLIANCE = "eeoc_compliance"
    REPORT_TYPE_CUSTOM_ANALYTICS = "custom_analytics"

    # Output format constants
    FORMAT_PDF = "pdf"
    FORMAT_EXCEL = "excel"
    FORMAT_CSV = "csv"

    # Page format constants
    PAGE_FORMAT_A4 = "A4"
    PAGE_FORMAT_LETTER = "Letter"

    # Content type mappings
    CONTENT_TYPES = {
        "pdf": "application/pdf",
        "excel": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "csv": "text/csv",
    }

    def __init__(
        self,
        db: AsyncSession,
        enabled: Optional[bool] = None,
        default_page_format: str = PAGE_FORMAT_A4,
    ) -> None:
        """
        Initialize the report template service.

        Args:
            db: Database session for data collection
            enabled: Whether report generation is enabled (default: True)
            default_page_format: Default page format for PDF reports (default: A4)
        """
        settings = get_settings()

        self.db = db
        self.enabled = enabled if enabled is not None else True
        self.default_page_format = default_page_format

        # Check for required libraries
        self._reportlab_available = False
        self._openpyxl_available = False

        if self.enabled:
            # Check for reportlab (PDF generation)
            try:
                import reportlab
                self._reportlab_available = True
                logger.info("ReportTemplateService initialized with reportlab support")
            except ImportError:
                logger.warning(
                    "reportlab not installed, PDF generation will be disabled. "
                    "Install with: pip install reportlab"
                )

            # Check for openpyxl (Excel generation)
            try:
                import openpyxl
                self._openpyxl_available = True
                logger.info("ReportTemplateService initialized with openpyxl support")
            except ImportError:
                logger.warning(
                    "openpyxl not installed, Excel generation will be disabled. "
                    "Install with: pip install openpyxl"
                )

        logger.info(
            f"ReportTemplateService initialized: "
            f"enabled={self.enabled}, "
            f"reportlab={self._reportlab_available}, "
            f"openpyxl={self._openpyxl_available}"
        )

    async def generate_report(
        self,
        config: ReportConfig,
        data: Optional[Dict[str, Any]] = None,
    ) -> ReportGenerationResult:
        """
        Generate a report based on configuration.

        This is the main entry point for report generation. It routes to the
        appropriate report type handler based on the configuration.

        Args:
            config: Report configuration
            data: Optional pre-collected data (if None, data will be collected)

        Returns:
            ReportGenerationResult with generated report or error

        Example:
            >>> config = ReportConfig(
            ...     report_type="candidate_summary",
            ...     output_format="pdf"
            ... )
            >>> result = await service.generate_report(config, {"resume_id": "123"})
        """
        if not self.enabled:
            return ReportGenerationResult(
                success=False,
                error_message="Report generation is disabled",
            )

        logger.info(
            f"Generating report: type={config.report_type}, "
            f"format={config.output_format}"
        )

        try:
            # Validate configuration
            self._validate_config(config)

            # Route to appropriate report type handler
            if config.report_type == self.REPORT_TYPE_CANDIDATE_SUMMARY:
                result = await self._generate_candidate_summary(config, data)
            elif config.report_type == self.REPORT_TYPE_HIRING_PIPELINE:
                result = await self._generate_hiring_pipeline(config, data)
            elif config.report_type == self.REPORT_TYPE_EEOC_COMPLIANCE:
                result = await self._generate_eeoc_compliance(config, data)
            elif config.report_type == self.REPORT_TYPE_CUSTOM_ANALYTICS:
                result = await self._generate_custom_analytics(config, data)
            else:
                return ReportGenerationResult(
                    success=False,
                    error_message=f"Unknown report type: {config.report_type}",
                )

            # Add metadata
            result.report_type = config.report_type
            result.generated_at = datetime.utcnow().isoformat() + "Z"

            if result.success:
                logger.info(
                    f"Report generated successfully: "
                    f"type={config.report_type}, "
                    f"size={result.file_size} bytes"
                )
            else:
                logger.error(
                    f"Report generation failed: "
                    f"type={config.report_type}, "
                    f"error={result.error_message}"
                )

            return result

        except Exception as e:
            logger.error(f"Error generating report: {e}", exc_info=True)
            return ReportGenerationResult(
                success=False,
                error_message=f"Report generation error: {str(e)}",
            )

    def _validate_config(self, config: ReportConfig) -> None:
        """
        Validate report configuration.

        Args:
            config: Report configuration to validate

        Raises:
            ValueError: If configuration is invalid
        """
        # Validate report type
        valid_types = [
            self.REPORT_TYPE_CANDIDATE_SUMMARY,
            self.REPORT_TYPE_HIRING_PIPELINE,
            self.REPORT_TYPE_EEOC_COMPLIANCE,
            self.REPORT_TYPE_CUSTOM_ANALYTICS,
        ]
        if config.report_type not in valid_types:
            raise ValueError(
                f"Invalid report type: {config.report_type}. "
                f"Must be one of: {', '.join(valid_types)}"
            )

        # Validate output format
        valid_formats = [self.FORMAT_PDF, self.FORMAT_EXCEL, self.FORMAT_CSV]
        if config.output_format not in valid_formats:
            raise ValueError(
                f"Invalid output format: {config.output_format}. "
                f"Must be one of: {', '.join(valid_formats)}"
            )

        # Check format-specific requirements
        if config.output_format == self.FORMAT_PDF and not self._reportlab_available:
            raise ValueError(
                "PDF generation requires reportlab library. "
                "Install with: pip install reportlab"
            )

        if config.output_format == self.FORMAT_EXCEL and not self._openpyxl_available:
            raise ValueError(
                "Excel generation requires openpyxl library. "
                "Install with: pip install openpyxl"
            )

    async def _generate_candidate_summary(
        self,
        config: ReportConfig,
        data: Optional[Dict[str, Any]] = None,
    ) -> ReportGenerationResult:
        """
        Generate candidate summary report.

        This report includes:
        - Resume metadata and contact information
        - Match scores and hiring stage
        - Skills and experience breakdown
        - Interview feedback and notes
        - Recommendation summary

        Args:
            config: Report configuration
            data: Pre-collected data or None to collect from database

        Returns:
            ReportGenerationResult with generated report or error

        Note:
            This is a placeholder implementation. The actual report generation
            logic will be implemented in subsequent subtasks.
        """
        logger.info("Generating candidate summary report")

        try:
            # Collect data if not provided
            if data is None:
                data = await self._collect_candidate_summary_data(config)

            # Validate required data
            if not data.get("resume_id"):
                raise ValueError("resume_id is required for candidate summary report")

            # Generate report based on output format
            if config.output_format == self.FORMAT_PDF:
                return await self._generate_candidate_summary_pdf(config, data)
            elif config.output_format == self.FORMAT_EXCEL:
                return await self._generate_candidate_summary_excel(config, data)
            elif config.output_format == self.FORMAT_CSV:
                return await self._generate_candidate_summary_csv(config, data)
            else:
                raise ValueError(f"Unsupported output format: {config.output_format}")

        except Exception as e:
            logger.error(f"Error generating candidate summary report: {e}", exc_info=True)
            return ReportGenerationResult(
                success=False,
                error_message=f"Failed to generate candidate summary report: {str(e)}",
            )

    async def _collect_candidate_summary_data(
        self, config: ReportConfig
    ) -> Dict[str, Any]:
        """
        Collect data for candidate summary report from database.

        Args:
            config: Report configuration with filters

        Returns:
            Dictionary with collected candidate data

        Raises:
            ValueError: If required filters are missing
        """
        from backend.models.resume import Resume
        from backend.models.candidate_rank import CandidateRank
        from backend.models.hiring_stage import HiringStage
        from backend.models.analysis_result import AnalysisResult

        # Extract resume_id from filters
        resume_id = config.filters.get("resume_id")
        if not resume_id:
            raise ValueError("resume_id is required in filters for candidate summary report")

        # Query resume data
        resume_stmt = select(Resume).where(Resume.id == resume_id)
        resume_result = await self.db.execute(resume_stmt)
        resume = resume_result.scalar_one_or_none()

        if not resume:
            raise ValueError(f"Resume not found: {resume_id}")

        # Query candidate rank data
        rank_stmt = select(CandidateRank).where(CandidateRank.resume_id == resume_id)
        if config.filters.get("vacancy_id"):
            rank_stmt = rank_stmt.where(CandidateRank.vacancy_id == config.filters["vacancy_id"])
        rank_result = await self.db.execute(rank_stmt)
        candidate_rank = rank_result.scalar_one_or_none()

        # Query hiring stage data
        stage_stmt = (
            select(HiringStage)
            .where(HiringStage.resume_id == resume_id)
            .order_by(HiringStage.created_at.desc())
        )
        stage_result = await self.db.execute(stage_stmt)
        hiring_stage = stage_result.scalar_one_or_none()

        # Query analysis result data
        analysis_stmt = select(AnalysisResult).where(AnalysisResult.resume_id == resume_id)
        analysis_result = await self.db.execute(analysis_stmt)
        analysis = analysis_result.scalar_one_or_none()

        # Build data dictionary
        data = {
            "resume_id": str(resume.id),
            "organization_id": resume.organization_id,
            "vacancy_id": str(resume.vacancy_id) if resume.vacancy_id else None,
            "filename": resume.filename,
            "status": resume.status.value if resume.status else None,
            "language": resume.language,
            "uploaded_at": resume.created_at.isoformat() if hasattr(resume, "created_at") else None,
        }

        # Add candidate rank data if available
        if candidate_rank:
            data["rank_score"] = float(candidate_rank.rank_score)
            data["rank_position"] = int(candidate_rank.rank_position) if candidate_rank.rank_position else None
            data["model_version"] = candidate_rank.model_version
            data["prediction_confidence"] = float(candidate_rank.prediction_confidence) if candidate_rank.prediction_confidence else None
            data["explanation_narrative"] = candidate_rank.explanation_narrative
            data["recommendation"] = candidate_rank.recommendation
            data["feature_contributions"] = candidate_rank.feature_contributions
            data["ranking_factors"] = candidate_rank.ranking_factors
            data["confidence_interval"] = candidate_rank.confidence_interval

        # Add hiring stage data if available
        if hiring_stage:
            data["hiring_stage"] = hiring_stage.stage_name.value if hiring_stage.stage_name else None
            data["stage_notes"] = hiring_stage.notes
            data["priority"] = hiring_stage.priority
            data["queue_entered_at"] = hiring_stage.queue_entered_at.isoformat() if hiring_stage.queue_entered_at else None

        # Add analysis result data if available
        if analysis:
            data["skills"] = analysis.skills or []
            data["experience_summary"] = analysis.experience_summary or {}
            data["keywords"] = analysis.keywords or []
            data["entities"] = analysis.entities or {}

        return data

    async def _generate_candidate_summary_pdf(
        self, config: ReportConfig, data: Dict[str, Any]
    ) -> ReportGenerationResult:
        """
        Generate candidate summary report in PDF format.

        Args:
            config: Report configuration
            data: Collected candidate data

        Returns:
            ReportGenerationResult with PDF bytes
        """
        from reportlab.lib.pagesizes import A4, Letter
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            SimpleDocTemplate,
            Paragraph,
            Spacer,
            PageBreak,
            Table,
            TableStyle,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        # Create PDF buffer
        buffer = io.BytesIO()

        # Get page size
        pagesize = A4 if config.page_format == self.PAGE_FORMAT_A4 else Letter

        # Create document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=pagesize,
            leftMargin=72,
            rightMargin=72,
            topMargin=72,
            bottomMargin=72,
        )

        # Get styles
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#2c3e50"),
            spaceAfter=30,
            alignment=TA_CENTER,
        )

        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=16,
            textColor=colors.HexColor("#34495e"),
            spaceAfter=12,
        )

        normal_style = styles["BodyText"]
        normal_style.fontSize = 11

        # Build story (content elements)
        story = []

        # Title
        title = config.title or "Candidate Summary Report"
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 0.3 * inch))

        # Resume metadata
        story.append(Paragraph("Resume Information", heading_style))

        metadata_data = [
            ["Resume ID:", data.get("resume_id", "N/A")],
            ["Filename:", data.get("filename", "N/A")],
            ["Status:", data.get("status", "N/A")],
            ["Language:", data.get("language", "N/A")],
        ]

        if data.get("uploaded_at"):
            metadata_data.append(["Uploaded:", data["uploaded_at"]])

        metadata_table = Table(metadata_data, colWidths=[2 * inch, 4 * inch])
        metadata_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 11),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(metadata_table)
        story.append(Spacer(1, 0.3 * inch))

        # Match scores and ranking
        if data.get("rank_score") is not None:
            story.append(Paragraph("Match Score & Ranking", heading_style))

            score_data = [
                ["Overall Match Score:", f"{data['rank_score']:.1%}"],
            ]

            if data.get("rank_position"):
                score_data.append(["Rank Position:", f"#{data['rank_position']}"])

            if data.get("prediction_confidence"):
                score_data.append(["Confidence:", f"{data['prediction_confidence']:.1%}"])

            if data.get("recommendation"):
                score_data.append(["Recommendation:", data["recommendation"].upper()])

            score_table = Table(score_data, colWidths=[2 * inch, 4 * inch])
            score_table.setStyle(TableStyle([
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("ALIGN", (1, 0), (1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 11),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(score_table)
            story.append(Spacer(1, 0.2 * inch))

            # Explanation narrative
            if data.get("explanation_narrative"):
                story.append(Paragraph("Explanation", heading_style))
                story.append(Paragraph(data["explanation_narrative"], normal_style))
                story.append(Spacer(1, 0.2 * inch))

        # Hiring stage
        if data.get("hiring_stage"):
            story.append(Paragraph("Hiring Stage", heading_style))

            stage_data = [
                ["Current Stage:", data["hiring_stage"].upper()],
            ]

            if data.get("priority"):
                stage_data.append(["Priority:", str(data["priority"])])

            if data.get("queue_entered_at"):
                stage_data.append(["Queue Entry:", data["queue_entered_at"]])

            stage_table = Table(stage_data, colWidths=[2 * inch, 4 * inch])
            stage_table.setStyle(TableStyle([
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                ("ALIGN", (1, 0), (1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 11),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]))
            story.append(stage_table)

            if data.get("stage_notes"):
                story.append(Spacer(1, 0.1 * inch))
                story.append(Paragraph(f"<b>Notes:</b> {data['stage_notes']}", normal_style))

            story.append(Spacer(1, 0.3 * inch))

        # Skills
        if data.get("skills"):
            story.append(Paragraph("Skills", heading_style))

            skills_list = []
            for skill in data["skills"]:
                if isinstance(skill, dict):
                    skill_name = skill.get("name", str(skill))
                else:
                    skill_name = str(skill)
                skills_list.append(skill_name)

            if skills_list:
                # Group skills into columns for better layout
                skills_text = ", ".join(skills_list)
                story.append(Paragraph(skills_text, normal_style))
                story.append(Spacer(1, 0.2 * inch))

        # Experience summary
        if data.get("experience_summary"):
            story.append(Paragraph("Experience Summary", heading_style))

            exp_summary = data["experience_summary"]
            exp_text = ""

            if "total_years_formatted" in exp_summary:
                exp_text += f"<b>Total Experience:</b> {exp_summary['total_years_formatted']}<br/>"
            elif "total_years" in exp_summary:
                exp_text += f"<b>Total Experience:</b> {exp_summary['total_years']:.1f} years<br/>"

            if "framework_specific" in exp_summary:
                exp_text += "<br/><b>Framework-Specific Experience:</b><br/>"
                for framework, experience in exp_summary["framework_specific"].items():
                    exp_text += f"• {framework}: {experience}<br/>"

            if exp_text:
                story.append(Paragraph(exp_text, normal_style))
                story.append(Spacer(1, 0.2 * inch))

        # Feature contributions (if available)
        if data.get("feature_contributions"):
            story.append(PageBreak())
            story.append(Paragraph("Feature Contribution Breakdown", heading_style))

            contributions = data["feature_contributions"]
            if isinstance(contributions, dict):
                # Convert dict to list format
                contrib_list = [
                    {"name": key, "contribution": value}
                    for key, value in contributions.items()
                ]
            else:
                contrib_list = contributions

            if contrib_list:
                table_data = [["Feature", "Contribution"]]

                for contrib in contrib_list[:10]:  # Limit to top 10
                    name = contrib.get("name", "N/A")
                    value = contrib.get("contribution", contrib.get("value", 0))

                    if isinstance(value, (int, float)):
                        value_str = f"{value:.3f}"
                    else:
                        value_str = str(value)

                    table_data.append([name, value_str])

                contrib_table = Table(table_data, colWidths=[3 * inch, 2 * inch])
                contrib_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#ecf0f1")),
                    ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#bdc3c7")),
                ]))

                story.append(contrib_table)
                story.append(Spacer(1, 0.2 * inch))

        # Confidence interval (if available)
        if data.get("confidence_interval"):
            story.append(Paragraph("Confidence Interval", heading_style))

            ci = data["confidence_interval"]
            ci_text = (
                f"<b>Lower Bound:</b> {ci.get('lower', 0):.1%}<br/>"
                f"<b>Upper Bound:</b> {ci.get('upper', 0):.1%}<br/>"
            )

            if "confidence_level" in ci:
                ci_text += f"<b>Confidence Level:</b> {ci.get('confidence_level', 0):.0%}<br/>"

            if "explanation" in ci:
                ci_text += f"<br/>{ci['explanation']}"

            story.append(Paragraph(ci_text, normal_style))
            story.append(Spacer(1, 0.2 * inch))

        # Footer
        story.append(Spacer(1, 0.5 * inch))
        footer_style = ParagraphStyle(
            "Footer",
            parent=normal_style,
            fontSize=9,
            textColor=colors.HexColor("#7f8c8d"),
            alignment=TA_CENTER,
        )
        timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
        footer_text = f"Generated: {timestamp} | Resume ID: {data['resume_id']}"
        if data.get("vacancy_id"):
            footer_text += f" | Vacancy ID: {data['vacancy_id']}"
        story.append(Paragraph(footer_text, footer_style))

        # Build PDF
        doc.build(story)

        # Get PDF bytes
        pdf_bytes = buffer.getvalue()
        buffer.close()

        # Generate filename
        filename = self._generate_filename(
            self.REPORT_TYPE_CANDIDATE_SUMMARY,
            self.FORMAT_PDF
        )

        return ReportGenerationResult(
            success=True,
            report_bytes=pdf_bytes,
            filename=filename,
            content_type=self.CONTENT_TYPES[self.FORMAT_PDF],
            file_size=len(pdf_bytes),
            metadata={
                "resume_id": data["resume_id"],
                "vacancy_id": data.get("vacancy_id"),
            },
        )

    async def _generate_candidate_summary_excel(
        self, config: ReportConfig, data: Dict[str, Any]
    ) -> ReportGenerationResult:
        """
        Generate candidate summary report in Excel format.

        Args:
            config: Report configuration
            data: Collected candidate data

        Returns:
            ReportGenerationResult with Excel bytes
        """
        # Placeholder for Excel generation
        # This would use openpyxl to create an Excel workbook
        return ReportGenerationResult(
            success=False,
            error_message="Excel format not yet implemented for candidate summary report",
        )

    async def _generate_candidate_summary_csv(
        self, config: ReportConfig, data: Dict[str, Any]
    ) -> ReportGenerationResult:
        """
        Generate candidate summary report in CSV format.

        Args:
            config: Report configuration
            data: Collected candidate data

        Returns:
            ReportGenerationResult with CSV bytes
        """
        # Placeholder for CSV generation
        # This would use csv module to create a CSV file
        return ReportGenerationResult(
            success=False,
            error_message="CSV format not yet implemented for candidate summary report",
        )

    async def _generate_hiring_pipeline(
        self,
        config: ReportConfig,
        data: Optional[Dict[str, Any]] = None,
    ) -> ReportGenerationResult:
        """
        Generate hiring pipeline report.

        This report includes:
        - Stage breakdowns with candidate counts
        - Time metrics (time in stage, total time to hire)
        - Conversion rates between stages
        - Pipeline health indicators
        - Bottleneck identification

        Args:
            config: Report configuration
            data: Pre-collected data or None to collect from database

        Returns:
            ReportGenerationResult with generated report or error
        """
        logger.info("Generating hiring pipeline report")

        try:
            # Collect data if not provided
            if data is None:
                data = await self._collect_hiring_pipeline_data(config)

            # Validate required data
            if not data.get("stages"):
                raise ValueError("No hiring stage data available for pipeline report")

            # Generate report based on output format
            if config.output_format == self.FORMAT_PDF:
                return await self._generate_hiring_pipeline_pdf(config, data)
            elif config.output_format == self.FORMAT_EXCEL:
                return await self._generate_hiring_pipeline_excel(config, data)
            elif config.output_format == self.FORMAT_CSV:
                return await self._generate_hiring_pipeline_csv(config, data)
            else:
                raise ValueError(f"Unsupported output format: {config.output_format}")

        except Exception as e:
            logger.error(f"Error generating hiring pipeline report: {e}", exc_info=True)
            return ReportGenerationResult(
                success=False,
                error_message=f"Failed to generate hiring pipeline report: {str(e)}",
            )

    async def _collect_hiring_pipeline_data(
        self, config: ReportConfig
    ) -> Dict[str, Any]:
        """
        Collect data for hiring pipeline report from database.

        This method aggregates hiring stage data to calculate:
        - Candidate counts per stage
        - Average time in each stage
        - Conversion rates between stages
        - Overall pipeline metrics

        Args:
            config: Report configuration with filters and date_range

        Returns:
            Dictionary with aggregated pipeline data

        Raises:
            ValueError: If required configuration is missing
        """
        from backend.models.hiring_stage import HiringStage, HiringStageName
        from backend.models.resume import Resume
        from sqlalchemy import func, and_

        # Build base query
        query = select(
            HiringStage.stage_name,
            func.count(HiringStage.id).label("count"),
            func.avg(
                func.extract(
                    "epoch",
                    func.coalesce(HiringStage.updated_at, HiringStage.created_at)
                    - HiringStage.queue_entered_at
                )
            ).label("avg_time_seconds"),
            func.min(HiringStage.created_at).label("first_entry"),
            func.max(HiringStage.created_at).label("last_entry"),
        ).group_by(HiringStage.stage_name)

        # Apply filters
        conditions = []

        # Organization filter
        if config.organization_id:
            query = query.join(Resume, HiringStage.resume_id == Resume.id)
            conditions.append(Resume.organization_id == config.organization_id)

        # Vacancy filter
        if config.filters.get("vacancy_id"):
            conditions.append(HiringStage.vacancy_id == config.filters["vacancy_id"])

        # Date range filter
        if config.date_range:
            if config.date_range.get("start"):
                start_date = datetime.fromisoformat(config.date_range["start"].replace("Z", "+00:00"))
                conditions.append(HiringStage.created_at >= start_date)
            if config.date_range.get("end"):
                end_date = datetime.fromisoformat(config.date_range["end"].replace("Z", "+00:00"))
                conditions.append(HiringStage.created_at <= end_date)

        if conditions:
            query = query.where(and_(*conditions))

        # Execute query
        result = await self.db.execute(query)
        stage_data = result.all()

        # Process stage data
        stages = []
        total_candidates = 0
        stage_counts = {}

        for row in stage_data:
            stage_name = row[0].value if hasattr(row[0], "value") else str(row[0])
            count = int(row[1])
            avg_time_seconds = float(row[2]) if row[2] else 0
            first_entry = row[3]
            last_entry = row[4]

            total_candidates += count
            stage_counts[stage_name] = count

            # Calculate average time in days
            avg_time_days = avg_time_seconds / 86400 if avg_time_seconds else 0

            stages.append({
                "stage_name": stage_name,
                "count": count,
                "avg_time_seconds": avg_time_seconds,
                "avg_time_days": avg_time_days,
                "avg_time_formatted": self._format_time_duration(avg_time_seconds),
                "first_entry": first_entry.isoformat() if first_entry else None,
                "last_entry": last_entry.isoformat() if last_entry else None,
            })

        # Calculate conversion rates
        conversion_rates = self._calculate_conversion_rates(stage_counts)

        # Calculate overall metrics
        total_time_query = select(
            func.avg(
                func.extract(
                    "epoch",
                    func.coalesce(HiringStage.updated_at, HiringStage.created_at)
                    - HiringStage.queue_entered_at
                )
            )
        )

        if conditions:
            if config.organization_id:
                total_time_query = total_time_query.join(Resume, HiringStage.resume_id == Resume.id)
            total_time_query = total_time_query.where(and_(*conditions))

        total_time_result = await self.db.execute(total_time_query)
        avg_total_time = total_time_result.scalar()
        avg_total_time_seconds = float(avg_total_time) if avg_total_time else 0

        # Build data dictionary
        data = {
            "stages": stages,
            "conversion_rates": conversion_rates,
            "total_candidates": total_candidates,
            "avg_time_to_hire_seconds": avg_total_time_seconds,
            "avg_time_to_hire_days": avg_total_time_seconds / 86400 if avg_total_time_seconds else 0,
            "avg_time_to_hire_formatted": self._format_time_duration(avg_total_time_seconds),
            "organization_id": config.organization_id,
            "vacancy_id": config.filters.get("vacancy_id"),
            "date_range": config.date_range,
        }

        # Calculate pipeline health indicators
        data["pipeline_health"] = self._calculate_pipeline_health(stages, conversion_rates)

        return data

    def _calculate_conversion_rates(self, stage_counts: Dict[str, int]) -> List[Dict[str, Any]]:
        """
        Calculate conversion rates between hiring stages.

        Args:
            stage_counts: Dictionary mapping stage names to candidate counts

        Returns:
            List of conversion rate data for each stage transition
        """
        from backend.models.hiring_stage import HiringStageName

        # Define stage order
        stage_order = [
            HiringStageName.APPLIED.value,
            HiringStageName.SCREENING.value,
            HiringStageName.INTERVIEW.value,
            HiringStageName.TECHNICAL.value,
            HiringStageName.OFFER.value,
            HiringStageName.HIRED.value,
        ]

        conversion_rates = []

        for i in range(len(stage_order) - 1):
            from_stage = stage_order[i]
            to_stage = stage_order[i + 1]

            from_count = stage_counts.get(from_stage, 0)
            to_count = stage_counts.get(to_stage, 0)

            if from_count > 0:
                rate = (to_count / from_count) * 100
            else:
                rate = 0

            conversion_rates.append({
                "from_stage": from_stage,
                "to_stage": to_stage,
                "from_count": from_count,
                "to_count": to_count,
                "conversion_rate": rate,
            })

        return conversion_rates

    def _calculate_pipeline_health(
        self, stages: List[Dict[str, Any]], conversion_rates: List[Dict[str, Any]]
    ) -> Dict[str, Any]:
        """
        Calculate pipeline health indicators and identify bottlenecks.

        Args:
            stages: List of stage data
            conversion_rates: List of conversion rate data

        Returns:
            Dictionary with pipeline health metrics
        """
        # Find bottlenecks (stages with lowest conversion rates)
        bottlenecks = []
        if conversion_rates:
            sorted_rates = sorted(conversion_rates, key=lambda x: x["conversion_rate"])
            # Consider stages with conversion rate < 50% as bottlenecks
            bottlenecks = [
                {
                    "stage": rate["from_stage"],
                    "conversion_rate": rate["conversion_rate"],
                }
                for rate in sorted_rates
                if rate["conversion_rate"] < 50 and rate["from_count"] > 0
            ]

        # Find stages with longest average time
        slow_stages = []
        if stages:
            sorted_stages = sorted(stages, key=lambda x: x["avg_time_days"], reverse=True)
            # Consider stages with > 7 days average time as slow
            slow_stages = [
                {
                    "stage": stage["stage_name"],
                    "avg_time_days": stage["avg_time_days"],
                    "avg_time_formatted": stage["avg_time_formatted"],
                }
                for stage in sorted_stages
                if stage["avg_time_days"] > 7
            ]

        # Calculate overall health score (0-100)
        # Based on conversion rates and time metrics
        health_score = 100

        # Penalize for bottlenecks
        if bottlenecks:
            health_score -= len(bottlenecks) * 10

        # Penalize for slow stages
        if slow_stages:
            health_score -= len(slow_stages) * 5

        # Ensure score is in valid range
        health_score = max(0, min(100, health_score))

        return {
            "score": health_score,
            "status": self._get_health_status(health_score),
            "bottlenecks": bottlenecks,
            "slow_stages": slow_stages,
        }

    def _get_health_status(self, score: int) -> str:
        """
        Get health status label based on score.

        Args:
            score: Health score (0-100)

        Returns:
            Status label (excellent, good, fair, poor)
        """
        if score >= 80:
            return "excellent"
        elif score >= 60:
            return "good"
        elif score >= 40:
            return "fair"
        else:
            return "poor"

    def _format_time_duration(self, seconds: float) -> str:
        """
        Format time duration in seconds to human-readable format.

        Args:
            seconds: Duration in seconds

        Returns:
            Formatted duration string (e.g., "5 days, 3 hours")
        """
        if seconds <= 0:
            return "N/A"

        days = int(seconds // 86400)
        remaining = seconds % 86400
        hours = int(remaining // 3600)
        remaining = remaining % 3600
        minutes = int(remaining // 60)

        parts = []
        if days > 0:
            parts.append(f"{days} day{'s' if days != 1 else ''}")
        if hours > 0:
            parts.append(f"{hours} hour{'s' if hours != 1 else ''}")
        if minutes > 0 and days == 0:  # Only show minutes if less than a day
            parts.append(f"{minutes} minute{'s' if minutes != 1 else ''}")

        if not parts:
            return "< 1 minute"

        return ", ".join(parts)

    async def _generate_hiring_pipeline_pdf(
        self, config: ReportConfig, data: Dict[str, Any]
    ) -> ReportGenerationResult:
        """
        Generate hiring pipeline report in PDF format.

        Args:
            config: Report configuration
            data: Collected pipeline data

        Returns:
            ReportGenerationResult with PDF bytes
        """
        from reportlab.lib.pagesizes import A4, Letter
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            SimpleDocTemplate,
            Paragraph,
            Spacer,
            PageBreak,
            Table,
            TableStyle,
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        # Create PDF buffer
        buffer = io.BytesIO()

        # Get page size
        pagesize = A4 if config.page_format == self.PAGE_FORMAT_A4 else Letter

        # Create document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=pagesize,
            leftMargin=72,
            rightMargin=72,
            topMargin=72,
            bottomMargin=72,
        )

        # Get styles
        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#2c3e50"),
            spaceAfter=30,
            alignment=TA_CENTER,
        )

        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=16,
            textColor=colors.HexColor("#34495e"),
            spaceAfter=12,
        )

        subheading_style = ParagraphStyle(
            "CustomSubHeading",
            parent=styles["Heading3"],
            fontSize=14,
            textColor=colors.HexColor("#34495e"),
            spaceAfter=10,
        )

        normal_style = styles["BodyText"]
        normal_style.fontSize = 11

        # Build story (content elements)
        story = []

        # Title
        title = config.title or "Hiring Pipeline Report"
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 0.3 * inch))

        # Summary metrics
        story.append(Paragraph("Pipeline Summary", heading_style))

        summary_data = [
            ["Total Candidates:", str(data["total_candidates"])],
            ["Average Time to Hire:", data["avg_time_to_hire_formatted"]],
            ["Pipeline Health:", f"{data['pipeline_health']['status'].upper()} ({data['pipeline_health']['score']}/100)"],
        ]

        if data.get("date_range"):
            date_range = data["date_range"]
            date_range_str = f"{date_range.get('start', 'N/A')} to {date_range.get('end', 'N/A')}"
            summary_data.append(["Date Range:", date_range_str])

        summary_table = Table(summary_data, colWidths=[2.5 * inch, 3.5 * inch])
        summary_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("ALIGN", (1, 0), (1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 11),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#ecf0f1")),
            ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#bdc3c7")),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.4 * inch))

        # Stage breakdown
        story.append(Paragraph("Stage Breakdown", heading_style))

        if data["stages"]:
            stage_table_data = [["Stage", "Candidates", "Avg. Time in Stage"]]

            for stage in data["stages"]:
                stage_table_data.append([
                    stage["stage_name"].upper(),
                    str(stage["count"]),
                    stage["avg_time_formatted"],
                ])

            stage_table = Table(stage_table_data, colWidths=[2 * inch, 1.5 * inch, 2.5 * inch])
            stage_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#ecf0f1")),
                ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#bdc3c7")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#ecf0f1"), colors.white]),
            ]))

            story.append(stage_table)
            story.append(Spacer(1, 0.4 * inch))

        # Conversion rates
        story.append(Paragraph("Conversion Rates", heading_style))

        if data["conversion_rates"]:
            conv_table_data = [["From Stage", "To Stage", "Candidates", "Conversion Rate"]]

            for rate in data["conversion_rates"]:
                conv_table_data.append([
                    rate["from_stage"].upper(),
                    rate["to_stage"].upper(),
                    f"{rate['to_count']} / {rate['from_count']}",
                    f"{rate['conversion_rate']:.1f}%",
                ])

            conv_table = Table(conv_table_data, colWidths=[1.8 * inch, 1.8 * inch, 1.2 * inch, 1.2 * inch])
            conv_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#ecf0f1")),
                ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#bdc3c7")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#ecf0f1"), colors.white]),
            ]))

            story.append(conv_table)
            story.append(Spacer(1, 0.4 * inch))

        # Pipeline health
        story.append(Paragraph("Pipeline Health Analysis", heading_style))

        health = data["pipeline_health"]
        health_text = f"<b>Overall Health Score:</b> {health['score']}/100 ({health['status'].upper()})<br/><br/>"

        if health["bottlenecks"]:
            health_text += "<b>Identified Bottlenecks:</b><br/>"
            for bottleneck in health["bottlenecks"]:
                health_text += f"• {bottleneck['stage'].upper()}: {bottleneck['conversion_rate']:.1f}% conversion rate<br/>"
            health_text += "<br/>"

        if health["slow_stages"]:
            health_text += "<b>Slow Stages:</b><br/>"
            for slow_stage in health["slow_stages"]:
                health_text += f"• {slow_stage['stage'].upper()}: {slow_stage['avg_time_formatted']} average time<br/>"

        if not health["bottlenecks"] and not health["slow_stages"]:
            health_text += "No significant issues identified. Pipeline is performing well."

        story.append(Paragraph(health_text, normal_style))
        story.append(Spacer(1, 0.3 * inch))

        # Footer
        story.append(Spacer(1, 0.5 * inch))
        footer_style = ParagraphStyle(
            "Footer",
            parent=normal_style,
            fontSize=9,
            textColor=colors.HexColor("#7f8c8d"),
            alignment=TA_CENTER,
        )
        timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
        footer_text = f"Generated: {timestamp}"
        if data.get("organization_id"):
            footer_text += f" | Organization: {data['organization_id']}"
        if data.get("vacancy_id"):
            footer_text += f" | Vacancy: {data['vacancy_id']}"
        story.append(Paragraph(footer_text, footer_style))

        # Build PDF
        doc.build(story)

        # Get PDF bytes
        pdf_bytes = buffer.getvalue()
        buffer.close()

        # Generate filename
        filename = self._generate_filename(
            self.REPORT_TYPE_HIRING_PIPELINE,
            self.FORMAT_PDF
        )

        return ReportGenerationResult(
            success=True,
            report_bytes=pdf_bytes,
            filename=filename,
            content_type=self.CONTENT_TYPES[self.FORMAT_PDF],
            file_size=len(pdf_bytes),
            metadata={
                "total_candidates": data["total_candidates"],
                "pipeline_health_score": data["pipeline_health"]["score"],
                "organization_id": data.get("organization_id"),
                "vacancy_id": data.get("vacancy_id"),
            },
        )

    async def _generate_hiring_pipeline_excel(
        self, config: ReportConfig, data: Dict[str, Any]
    ) -> ReportGenerationResult:
        """
        Generate hiring pipeline report in Excel format.

        Args:
            config: Report configuration
            data: Collected pipeline data

        Returns:
            ReportGenerationResult with Excel bytes
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Hiring Pipeline Report"])
        ws_summary.append([])
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Total Candidates", data["total_candidates"]])
        ws_summary.append(["Average Time to Hire", data["avg_time_to_hire_formatted"]])
        ws_summary.append(["Pipeline Health Score", f"{data['pipeline_health']['score']}/100"])
        ws_summary.append(["Pipeline Health Status", data["pipeline_health"]["status"].upper()])

        if data.get("date_range"):
            ws_summary.append(["Date Range Start", data["date_range"].get("start", "N/A")])
            ws_summary.append(["Date Range End", data["date_range"].get("end", "N/A")])

        # Style summary sheet
        ws_summary["A1"].font = Font(size=16, bold=True)
        ws_summary["A3"].font = Font(bold=True)
        ws_summary["B3"].font = Font(bold=True)
        ws_summary["A3"].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        ws_summary["B3"].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        ws_summary["A3"].font = Font(color="FFFFFF", bold=True)
        ws_summary["B3"].font = Font(color="FFFFFF", bold=True)

        # Adjust column widths
        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 40

        # Create Stage Breakdown sheet
        ws_stages = wb.create_sheet("Stage Breakdown")
        ws_stages.append(["Stage", "Candidate Count", "Avg Time (Days)", "Avg Time (Formatted)", "First Entry", "Last Entry"])

        for stage in data["stages"]:
            ws_stages.append([
                stage["stage_name"].upper(),
                stage["count"],
                round(stage["avg_time_days"], 2),
                stage["avg_time_formatted"],
                stage["first_entry"],
                stage["last_entry"],
            ])

        # Style stages sheet
        for cell in ws_stages[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Adjust column widths
        ws_stages.column_dimensions["A"].width = 20
        ws_stages.column_dimensions["B"].width = 18
        ws_stages.column_dimensions["C"].width = 18
        ws_stages.column_dimensions["D"].width = 25
        ws_stages.column_dimensions["E"].width = 22
        ws_stages.column_dimensions["F"].width = 22

        # Create Conversion Rates sheet
        ws_conversion = wb.create_sheet("Conversion Rates")
        ws_conversion.append(["From Stage", "To Stage", "From Count", "To Count", "Conversion Rate (%)"])

        for rate in data["conversion_rates"]:
            ws_conversion.append([
                rate["from_stage"].upper(),
                rate["to_stage"].upper(),
                rate["from_count"],
                rate["to_count"],
                round(rate["conversion_rate"], 2),
            ])

        # Style conversion sheet
        for cell in ws_conversion[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Adjust column widths
        ws_conversion.column_dimensions["A"].width = 20
        ws_conversion.column_dimensions["B"].width = 20
        ws_conversion.column_dimensions["C"].width = 15
        ws_conversion.column_dimensions["D"].width = 15
        ws_conversion.column_dimensions["E"].width = 22

        # Create Pipeline Health sheet
        ws_health = wb.create_sheet("Pipeline Health")
        ws_health.append(["Pipeline Health Analysis"])
        ws_health.append([])
        ws_health.append(["Health Score", data["pipeline_health"]["score"]])
        ws_health.append(["Health Status", data["pipeline_health"]["status"].upper()])
        ws_health.append([])

        if data["pipeline_health"]["bottlenecks"]:
            ws_health.append(["Bottlenecks"])
            ws_health.append(["Stage", "Conversion Rate (%)"])
            for bottleneck in data["pipeline_health"]["bottlenecks"]:
                ws_health.append([
                    bottleneck["stage"].upper(),
                    round(bottleneck["conversion_rate"], 2),
                ])
            ws_health.append([])

        if data["pipeline_health"]["slow_stages"]:
            ws_health.append(["Slow Stages"])
            ws_health.append(["Stage", "Avg Time (Days)", "Avg Time (Formatted)"])
            for slow_stage in data["pipeline_health"]["slow_stages"]:
                ws_health.append([
                    slow_stage["stage"].upper(),
                    round(slow_stage["avg_time_days"], 2),
                    slow_stage["avg_time_formatted"],
                ])

        # Style health sheet
        ws_health["A1"].font = Font(size=14, bold=True)
        ws_health.column_dimensions["A"].width = 25
        ws_health.column_dimensions["B"].width = 25
        ws_health.column_dimensions["C"].width = 25

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()

        # Generate filename
        filename = self._generate_filename(
            self.REPORT_TYPE_HIRING_PIPELINE,
            self.FORMAT_EXCEL
        )

        return ReportGenerationResult(
            success=True,
            report_bytes=excel_bytes,
            filename=filename,
            content_type=self.CONTENT_TYPES[self.FORMAT_EXCEL],
            file_size=len(excel_bytes),
            metadata={
                "total_candidates": data["total_candidates"],
                "pipeline_health_score": data["pipeline_health"]["score"],
                "organization_id": data.get("organization_id"),
                "vacancy_id": data.get("vacancy_id"),
            },
        )

    async def _generate_hiring_pipeline_csv(
        self, config: ReportConfig, data: Dict[str, Any]
    ) -> ReportGenerationResult:
        """
        Generate hiring pipeline report in CSV format.

        Args:
            config: Report configuration
            data: Collected pipeline data

        Returns:
            ReportGenerationResult with CSV bytes
        """
        import csv

        # Create CSV buffer
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        # Write summary section
        writer.writerow(["Hiring Pipeline Report"])
        writer.writerow([])
        writer.writerow(["Summary Metrics"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Candidates", data["total_candidates"]])
        writer.writerow(["Average Time to Hire", data["avg_time_to_hire_formatted"]])
        writer.writerow(["Pipeline Health Score", f"{data['pipeline_health']['score']}/100"])
        writer.writerow(["Pipeline Health Status", data["pipeline_health"]["status"].upper()])

        if data.get("date_range"):
            writer.writerow(["Date Range Start", data["date_range"].get("start", "N/A")])
            writer.writerow(["Date Range End", data["date_range"].get("end", "N/A")])

        writer.writerow([])
        writer.writerow([])

        # Write stage breakdown section
        writer.writerow(["Stage Breakdown"])
        writer.writerow(["Stage", "Candidate Count", "Avg Time (Days)", "Avg Time (Formatted)", "First Entry", "Last Entry"])

        for stage in data["stages"]:
            writer.writerow([
                stage["stage_name"].upper(),
                stage["count"],
                round(stage["avg_time_days"], 2),
                stage["avg_time_formatted"],
                stage["first_entry"],
                stage["last_entry"],
            ])

        writer.writerow([])
        writer.writerow([])

        # Write conversion rates section
        writer.writerow(["Conversion Rates"])
        writer.writerow(["From Stage", "To Stage", "From Count", "To Count", "Conversion Rate (%)"])

        for rate in data["conversion_rates"]:
            writer.writerow([
                rate["from_stage"].upper(),
                rate["to_stage"].upper(),
                rate["from_count"],
                rate["to_count"],
                round(rate["conversion_rate"], 2),
            ])

        writer.writerow([])
        writer.writerow([])

        # Write pipeline health section
        writer.writerow(["Pipeline Health Analysis"])
        writer.writerow(["Health Score", data["pipeline_health"]["score"]])
        writer.writerow(["Health Status", data["pipeline_health"]["status"].upper()])
        writer.writerow([])

        if data["pipeline_health"]["bottlenecks"]:
            writer.writerow(["Bottlenecks"])
            writer.writerow(["Stage", "Conversion Rate (%)"])
            for bottleneck in data["pipeline_health"]["bottlenecks"]:
                writer.writerow([
                    bottleneck["stage"].upper(),
                    round(bottleneck["conversion_rate"], 2),
                ])
            writer.writerow([])

        if data["pipeline_health"]["slow_stages"]:
            writer.writerow(["Slow Stages"])
            writer.writerow(["Stage", "Avg Time (Days)", "Avg Time (Formatted)"])
            for slow_stage in data["pipeline_health"]["slow_stages"]:
                writer.writerow([
                    slow_stage["stage"].upper(),
                    round(slow_stage["avg_time_days"], 2),
                    slow_stage["avg_time_formatted"],
                ])

        # Get CSV content
        csv_content = buffer.getvalue()
        buffer.close()

        # Convert to bytes
        csv_bytes = csv_content.encode("utf-8")

        # Generate filename
        filename = self._generate_filename(
            self.REPORT_TYPE_HIRING_PIPELINE,
            self.FORMAT_CSV
        )

        return ReportGenerationResult(
            success=True,
            report_bytes=csv_bytes,
            filename=filename,
            content_type=self.CONTENT_TYPES[self.FORMAT_CSV],
            file_size=len(csv_bytes),
            metadata={
                "total_candidates": data["total_candidates"],
                "pipeline_health_score": data["pipeline_health"]["score"],
                "organization_id": data.get("organization_id"),
                "vacancy_id": data.get("vacancy_id"),
            },
        )

    async def _generate_eeoc_compliance(
        self,
        config: ReportConfig,
        data: Optional[Dict[str, Any]] = None,
    ) -> ReportGenerationResult:
        """
        Generate EEOC compliance report.

        This report includes:
        - Diversity metrics by demographic category (gender, age, ethnicity)
        - Hiring statistics by protected class
        - Demographic breakdowns with counts and percentages
        - Sample sizes and confidence indicators
        - Date range and filtering details

        Args:
            config: Report configuration
            data: Pre-collected data or None to collect from database

        Returns:
            ReportGenerationResult with generated report or error
        """
        logger.info("Generating EEOC compliance report")

        try:
            # Collect data if not provided
            if data is None:
                data = await self._collect_eeoc_compliance_data(config)

            # Validate required data
            if not data.get("total_candidates"):
                raise ValueError("No candidate data available for EEOC compliance report")

            # Generate report based on output format
            if config.output_format == self.FORMAT_PDF:
                return await self._generate_eeoc_compliance_pdf(config, data)
            elif config.output_format == self.FORMAT_EXCEL:
                return await self._generate_eeoc_compliance_excel(config, data)
            elif config.output_format == self.FORMAT_CSV:
                return await self._generate_eeoc_compliance_csv(config, data)
            else:
                raise ValueError(f"Unsupported output format: {config.output_format}")

        except Exception as e:
            logger.error(f"Error generating EEOC compliance report: {e}", exc_info=True)
            return ReportGenerationResult(
                success=False,
                error_message=f"Failed to generate EEOC compliance report: {str(e)}",
            )

    async def _generate_custom_analytics(
        self,
        config: ReportConfig,
        data: Optional[Dict[str, Any]] = None,
    ) -> ReportGenerationResult:
        """
        Generate custom analytics report.

        This report includes:
        - User-selected metrics and dimensions
        - Custom filters and date ranges
        - Flexible visualizations
        - Exportable data tables
        - Summary statistics

        Args:
            config: Report configuration
            data: Pre-collected data or None to collect from database

        Returns:
            ReportGenerationResult with generated report or error
        """
        try:
            logger.info(f"Generating custom analytics report in {config.output_format} format")

            # Collect data if not provided
            if data is None:
                data = await self._collect_custom_analytics_data(config)

            # Generate report based on output format
            if config.output_format == self.FORMAT_PDF:
                result = await self._generate_custom_analytics_pdf(config, data)
            elif config.output_format == self.FORMAT_EXCEL:
                result = await self._generate_custom_analytics_excel(config, data)
            elif config.output_format == self.FORMAT_CSV:
                result = await self._generate_custom_analytics_csv(config, data)
            else:
                return ReportGenerationResult(
                    success=False,
                    error_message=f"Unsupported output format: {config.output_format}",
                )

            # Add report type and generation timestamp
            result.report_type = self.REPORT_TYPE_CUSTOM_ANALYTICS
            result.generated_at = datetime.utcnow().isoformat() + "Z"

            logger.info(f"Custom analytics report generated successfully: {result.filename}")
            return result

        except Exception as e:
            logger.error(f"Error generating custom analytics report: {e}", exc_info=True)
            return ReportGenerationResult(
                success=False,
                error_message=f"Failed to generate custom analytics report: {str(e)}",
            )

    async def _collect_custom_analytics_data(
        self, config: ReportConfig
    ) -> Dict[str, Any]:
        """
        Collect data for custom analytics report from database.

        This method aggregates data based on user-selected metrics:
        - Candidate counts and distributions
        - Hiring stage metrics
        - Time to hire statistics
        - Conversion rates
        - Skills and experience distributions
        - Source/channel metrics

        Args:
            config: Report configuration with metrics, filters, and date_range

        Returns:
            Dictionary with aggregated custom analytics data

        Raises:
            ValueError: If required configuration is missing
        """
        from backend.models.resume import Resume
        from backend.models.hiring_stage import HiringStage, HiringStageName
        from backend.models.parsed_resume import ParsedResume
        from sqlalchemy import func, and_

        # Initialize data structure
        analytics_data: Dict[str, Any] = {
            "metrics": config.metrics or [],
            "filters": config.filters,
            "date_range": config.date_range,
            "organization_id": config.organization_id,
            "total_candidates": 0,
            "summary_statistics": {},
            "metric_data": {},
        }

        # Build base query conditions
        conditions = []

        # Organization filter
        if config.organization_id:
            conditions.append(Resume.organization_id == config.organization_id)

        # Date range filter
        if config.date_range:
            if config.date_range.get("start"):
                start_date = datetime.fromisoformat(config.date_range["start"].replace("Z", "+00:00"))
                conditions.append(Resume.created_at >= start_date)
            if config.date_range.get("end"):
                end_date = datetime.fromisoformat(config.date_range["end"].replace("Z", "+00:00"))
                conditions.append(Resume.created_at <= end_date)

        # Additional custom filters
        if config.filters:
            # Status filter
            if config.filters.get("status"):
                conditions.append(Resume.status == config.filters["status"])

            # Vacancy filter
            if config.filters.get("vacancy_id"):
                # Will be joined with HiringStage later
                pass

        # Collect total candidates
        total_query = select(func.count(Resume.id))
        if conditions:
            total_query = total_query.where(and_(*conditions))

        total_result = await self.db.execute(total_query)
        analytics_data["total_candidates"] = total_result.scalar() or 0

        # Collect metric-specific data based on requested metrics
        metrics = config.metrics if config.metrics else [
            "candidates_by_stage",
            "candidates_by_status",
            "time_to_hire",
            "conversion_rates",
            "skills_distribution",
        ]

        # Metric: Candidates by stage
        if "candidates_by_stage" in metrics:
            stage_query = select(
                HiringStage.stage_name,
                func.count(HiringStage.id).label("count")
            ).select_from(HiringStage)

            if conditions:
                stage_query = stage_query.join(Resume, HiringStage.resume_id == Resume.id).where(and_(*conditions))

            if config.filters.get("vacancy_id"):
                stage_query = stage_query.where(HiringStage.vacancy_id == config.filters["vacancy_id"])

            stage_query = stage_query.group_by(HiringStage.stage_name)

            stage_result = await self.db.execute(stage_query)
            stage_data = stage_result.all()

            stage_breakdown = {}
            for row in stage_data:
                stage_name = row[0].value if hasattr(row[0], "value") else str(row[0])
                count = int(row[1])
                stage_breakdown[stage_name] = {
                    "count": count,
                    "percentage": round((count / analytics_data["total_candidates"] * 100), 2) if analytics_data["total_candidates"] > 0 else 0,
                }

            analytics_data["metric_data"]["candidates_by_stage"] = stage_breakdown

        # Metric: Candidates by status
        if "candidates_by_status" in metrics:
            status_query = select(
                Resume.status,
                func.count(Resume.id).label("count")
            ).group_by(Resume.status)

            if conditions:
                status_query = status_query.where(and_(*conditions))

            status_result = await self.db.execute(status_query)
            status_data = status_result.all()

            status_breakdown = {}
            for row in status_data:
                status = row[0].value if hasattr(row[0], "value") else str(row[0]) if row[0] else "unknown"
                count = int(row[1])
                status_breakdown[status] = {
                    "count": count,
                    "percentage": round((count / analytics_data["total_candidates"] * 100), 2) if analytics_data["total_candidates"] > 0 else 0,
                }

            analytics_data["metric_data"]["candidates_by_status"] = status_breakdown

        # Metric: Time to hire
        if "time_to_hire" in metrics:
            time_query = select(
                func.avg(
                    func.extract(
                        "epoch",
                        func.coalesce(HiringStage.updated_at, HiringStage.created_at)
                        - HiringStage.queue_entered_at
                    )
                ).label("avg_time_seconds"),
                func.min(
                    func.extract(
                        "epoch",
                        func.coalesce(HiringStage.updated_at, HiringStage.created_at)
                        - HiringStage.queue_entered_at
                    )
                ).label("min_time_seconds"),
                func.max(
                    func.extract(
                        "epoch",
                        func.coalesce(HiringStage.updated_at, HiringStage.created_at)
                        - HiringStage.queue_entered_at
                    )
                ).label("max_time_seconds"),
            ).select_from(HiringStage)

            if conditions:
                time_query = time_query.join(Resume, HiringStage.resume_id == Resume.id).where(and_(*conditions))

            if config.filters.get("vacancy_id"):
                time_query = time_query.where(HiringStage.vacancy_id == config.filters["vacancy_id"])

            time_result = await self.db.execute(time_query)
            time_data = time_result.one_or_none()

            if time_data:
                avg_time_seconds = float(time_data[0]) if time_data[0] else 0
                min_time_seconds = float(time_data[1]) if time_data[1] else 0
                max_time_seconds = float(time_data[2]) if time_data[2] else 0

                analytics_data["metric_data"]["time_to_hire"] = {
                    "avg_time_seconds": avg_time_seconds,
                    "avg_time_days": avg_time_seconds / 86400 if avg_time_seconds else 0,
                    "avg_time_formatted": self._format_time_duration(avg_time_seconds),
                    "min_time_seconds": min_time_seconds,
                    "min_time_days": min_time_seconds / 86400 if min_time_seconds else 0,
                    "min_time_formatted": self._format_time_duration(min_time_seconds),
                    "max_time_seconds": max_time_seconds,
                    "max_time_days": max_time_seconds / 86400 if max_time_seconds else 0,
                    "max_time_formatted": self._format_time_duration(max_time_seconds),
                }
            else:
                analytics_data["metric_data"]["time_to_hire"] = {
                    "avg_time_seconds": 0,
                    "avg_time_days": 0,
                    "avg_time_formatted": "N/A",
                    "min_time_seconds": 0,
                    "min_time_days": 0,
                    "min_time_formatted": "N/A",
                    "max_time_seconds": 0,
                    "max_time_days": 0,
                    "max_time_formatted": "N/A",
                }

        # Metric: Conversion rates
        if "conversion_rates" in metrics:
            # Collect stage counts for conversion calculation
            stage_query = select(
                HiringStage.stage_name,
                func.count(HiringStage.id).label("count")
            ).select_from(HiringStage)

            if conditions:
                stage_query = stage_query.join(Resume, HiringStage.resume_id == Resume.id).where(and_(*conditions))

            if config.filters.get("vacancy_id"):
                stage_query = stage_query.where(HiringStage.vacancy_id == config.filters["vacancy_id"])

            stage_query = stage_query.group_by(HiringStage.stage_name)

            stage_result = await self.db.execute(stage_query)
            stage_data = stage_result.all()

            stage_counts = {}
            for row in stage_data:
                stage_name = row[0].value if hasattr(row[0], "value") else str(row[0])
                stage_counts[stage_name] = int(row[1])

            # Calculate conversion rates
            conversion_rates = self._calculate_conversion_rates(stage_counts)
            analytics_data["metric_data"]["conversion_rates"] = conversion_rates

        # Metric: Skills distribution
        if "skills_distribution" in metrics:
            skills_query = select(
                ParsedResume.skills,
                func.count(ParsedResume.id).label("count")
            ).select_from(ParsedResume)

            if conditions:
                skills_query = skills_query.join(Resume, ParsedResume.resume_id == Resume.id).where(and_(*conditions))

            skills_query = skills_query.where(ParsedResume.skills.isnot(None))

            skills_result = await self.db.execute(skills_query)
            skills_data = skills_result.all()

            # Aggregate skills
            skills_count = {}
            for row in skills_data:
                skills_list = row[0] if isinstance(row[0], list) else []
                for skill in skills_list:
                    if skill:
                        skill = str(skill).strip()
                        skills_count[skill] = skills_count.get(skill, 0) + 1

            # Sort by count and take top 20
            top_skills = sorted(skills_count.items(), key=lambda x: x[1], reverse=True)[:20]

            analytics_data["metric_data"]["skills_distribution"] = {
                skill: {
                    "count": count,
                    "percentage": round((count / analytics_data["total_candidates"] * 100), 2) if analytics_data["total_candidates"] > 0 else 0,
                }
                for skill, count in top_skills
            }

        # Calculate summary statistics
        analytics_data["summary_statistics"] = {
            "total_candidates": analytics_data["total_candidates"],
            "total_metrics": len(metrics),
            "report_generated_at": datetime.utcnow().isoformat() + "Z",
        }

        logger.info(
            f"Collected custom analytics data: {analytics_data['total_candidates']} candidates, "
            f"{len(metrics)} metrics"
        )

        return analytics_data

    async def _generate_custom_analytics_pdf(
        self, config: ReportConfig, data: Dict[str, Any]
    ) -> ReportGenerationResult:
        """
        Generate custom analytics report in PDF format.

        Args:
            config: Report configuration
            data: Collected custom analytics data

        Returns:
            ReportGenerationResult with PDF bytes
        """
        if not self._reportlab_available:
            return ReportGenerationResult(
                success=False,
                error_message="PDF generation requires reportlab library",
            )

        from reportlab.lib.pagesizes import A4, letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

        # Create PDF buffer
        buffer = io.BytesIO()

        # Determine page size
        pagesize = letter if config.page_format == self.PAGE_FORMAT_LETTER else A4

        # Create document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=pagesize,
            rightMargin=0.75 * inch,
            leftMargin=0.75 * inch,
            topMargin=1.0 * inch,
            bottomMargin=0.75 * inch,
        )

        # Container for PDF elements
        story = []

        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#2c3e50"),
            spaceAfter=30,
            alignment=TA_CENTER,
        )
        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=16,
            textColor=colors.HexColor("#34495e"),
            spaceAfter=12,
            spaceBefore=12,
        )
        normal_style = styles["Normal"]

        # Title
        title = config.title or "Custom Analytics Report"
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 0.3 * inch))

        # Summary section
        story.append(Paragraph("Summary Statistics", heading_style))

        summary_data = [
            ["Metric", "Value"],
            ["Total Candidates", str(data["total_candidates"])],
            ["Metrics Analyzed", str(len(data["metrics"]))],
            ["Report Generated", data["summary_statistics"]["report_generated_at"]],
        ]

        if data.get("date_range"):
            summary_data.append(["Date Range Start", data["date_range"].get("start", "N/A")])
            summary_data.append(["Date Range End", data["date_range"].get("end", "N/A")])

        if data.get("organization_id"):
            summary_data.append(["Organization ID", data["organization_id"]])

        summary_table = Table(summary_data, colWidths=[3 * inch, 4 * inch])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 12),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3 * inch))

        # Metric-specific sections
        metric_data = data.get("metric_data", {})

        # Candidates by stage
        if "candidates_by_stage" in metric_data:
            story.append(Paragraph("Candidates by Hiring Stage", heading_style))

            stage_data = metric_data["candidates_by_stage"]
            table_data = [["Stage", "Count", "Percentage"]]

            for stage, values in sorted(stage_data.items()):
                table_data.append([
                    stage.upper(),
                    str(values["count"]),
                    f"{values['percentage']}%",
                ])

            stage_table = Table(table_data, colWidths=[3 * inch, 2 * inch, 2 * inch])
            stage_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(stage_table)
            story.append(Spacer(1, 0.3 * inch))

        # Candidates by status
        if "candidates_by_status" in metric_data:
            story.append(Paragraph("Candidates by Status", heading_style))

            status_data = metric_data["candidates_by_status"]
            table_data = [["Status", "Count", "Percentage"]]

            for status, values in sorted(status_data.items()):
                table_data.append([
                    status.upper(),
                    str(values["count"]),
                    f"{values['percentage']}%",
                ])

            status_table = Table(table_data, colWidths=[3 * inch, 2 * inch, 2 * inch])
            status_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(status_table)
            story.append(Spacer(1, 0.3 * inch))

        # Time to hire
        if "time_to_hire" in metric_data:
            story.append(Paragraph("Time to Hire Metrics", heading_style))

            time_data = metric_data["time_to_hire"]
            table_data = [
                ["Metric", "Value"],
                ["Average Time", time_data["avg_time_formatted"]],
                ["Minimum Time", time_data["min_time_formatted"]],
                ["Maximum Time", time_data["max_time_formatted"]],
            ]

            time_table = Table(table_data, colWidths=[3 * inch, 4 * inch])
            time_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(time_table)
            story.append(Spacer(1, 0.3 * inch))

        # Conversion rates
        if "conversion_rates" in metric_data:
            story.append(Paragraph("Stage Conversion Rates", heading_style))

            conversion_data = metric_data["conversion_rates"]
            table_data = [["From Stage", "To Stage", "Conversion Rate"]]

            for conversion in conversion_data:
                table_data.append([
                    conversion["from_stage"].upper(),
                    conversion["to_stage"].upper(),
                    f"{round(conversion['conversion_rate'], 2)}%",
                ])

            conversion_table = Table(table_data, colWidths=[2.5 * inch, 2.5 * inch, 2 * inch])
            conversion_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(conversion_table)
            story.append(Spacer(1, 0.3 * inch))

        # Skills distribution
        if "skills_distribution" in metric_data:
            story.append(Paragraph("Top Skills Distribution", heading_style))

            skills_data = metric_data["skills_distribution"]
            table_data = [["Skill", "Count", "Percentage"]]

            for skill, values in list(skills_data.items())[:15]:  # Top 15 skills for PDF
                table_data.append([
                    skill,
                    str(values["count"]),
                    f"{values['percentage']}%",
                ])

            skills_table = Table(table_data, colWidths=[3 * inch, 2 * inch, 2 * inch])
            skills_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(skills_table)
            story.append(Spacer(1, 0.3 * inch))

        # Footer
        footer_style = ParagraphStyle(
            "Footer",
            parent=normal_style,
            fontSize=9,
            textColor=colors.HexColor("#7f8c8d"),
            alignment=TA_CENTER,
        )
        timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
        footer_text = f"Generated: {timestamp}"
        if data.get("organization_id"):
            footer_text += f" | Organization: {data['organization_id']}"
        story.append(Paragraph(footer_text, footer_style))

        # Build PDF
        doc.build(story)

        # Get PDF bytes
        pdf_bytes = buffer.getvalue()
        buffer.close()

        # Generate filename
        filename = self._generate_filename(
            self.REPORT_TYPE_CUSTOM_ANALYTICS,
            self.FORMAT_PDF
        )

        return ReportGenerationResult(
            success=True,
            report_bytes=pdf_bytes,
            filename=filename,
            content_type=self.CONTENT_TYPES[self.FORMAT_PDF],
            file_size=len(pdf_bytes),
            metadata={
                "total_candidates": data["total_candidates"],
                "metrics_count": len(data["metrics"]),
                "organization_id": data.get("organization_id"),
            },
        )

    async def _generate_custom_analytics_excel(
        self, config: ReportConfig, data: Dict[str, Any]
    ) -> ReportGenerationResult:
        """
        Generate custom analytics report in Excel format.

        Args:
            config: Report configuration
            data: Collected custom analytics data

        Returns:
            ReportGenerationResult with Excel bytes
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Custom Analytics Report"])
        ws_summary.append([])
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Total Candidates", data["total_candidates"]])
        ws_summary.append(["Metrics Analyzed", len(data["metrics"])])
        ws_summary.append(["Report Generated", data["summary_statistics"]["report_generated_at"]])

        if data.get("date_range"):
            ws_summary.append(["Date Range Start", data["date_range"].get("start", "N/A")])
            ws_summary.append(["Date Range End", data["date_range"].get("end", "N/A")])

        if data.get("organization_id"):
            ws_summary.append(["Organization ID", data["organization_id"]])

        # Style summary sheet
        ws_summary["A1"].font = Font(size=16, bold=True)
        ws_summary["A3"].font = Font(bold=True)
        ws_summary["B3"].font = Font(bold=True)
        ws_summary["A3"].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        ws_summary["B3"].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        ws_summary["A3"].font = Font(color="FFFFFF", bold=True)
        ws_summary["B3"].font = Font(color="FFFFFF", bold=True)

        # Adjust column widths
        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 40

        metric_data = data.get("metric_data", {})

        # Candidates by stage sheet
        if "candidates_by_stage" in metric_data:
            ws_stages = wb.create_sheet("Candidates by Stage")
            ws_stages.append(["Stage", "Count", "Percentage (%)"])

            stage_data = metric_data["candidates_by_stage"]
            for stage, values in sorted(stage_data.items()):
                ws_stages.append([
                    stage.upper(),
                    values["count"],
                    values["percentage"],
                ])

            # Style header
            for cell in ws_stages[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Adjust column widths
            ws_stages.column_dimensions["A"].width = 25
            ws_stages.column_dimensions["B"].width = 15
            ws_stages.column_dimensions["C"].width = 18

        # Candidates by status sheet
        if "candidates_by_status" in metric_data:
            ws_status = wb.create_sheet("Candidates by Status")
            ws_status.append(["Status", "Count", "Percentage (%)"])

            status_data = metric_data["candidates_by_status"]
            for status, values in sorted(status_data.items()):
                ws_status.append([
                    status.upper(),
                    values["count"],
                    values["percentage"],
                ])

            # Style header
            for cell in ws_status[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Adjust column widths
            ws_status.column_dimensions["A"].width = 25
            ws_status.column_dimensions["B"].width = 15
            ws_status.column_dimensions["C"].width = 18

        # Time to hire sheet
        if "time_to_hire" in metric_data:
            ws_time = wb.create_sheet("Time to Hire")
            ws_time.append(["Metric", "Value"])

            time_data = metric_data["time_to_hire"]
            ws_time.append(["Average Time", time_data["avg_time_formatted"]])
            ws_time.append(["Average Time (Days)", round(time_data["avg_time_days"], 2)])
            ws_time.append(["Minimum Time", time_data["min_time_formatted"]])
            ws_time.append(["Minimum Time (Days)", round(time_data["min_time_days"], 2)])
            ws_time.append(["Maximum Time", time_data["max_time_formatted"]])
            ws_time.append(["Maximum Time (Days)", round(time_data["max_time_days"], 2)])

            # Style header
            for cell in ws_time[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)

            # Adjust column widths
            ws_time.column_dimensions["A"].width = 30
            ws_time.column_dimensions["B"].width = 25

        # Conversion rates sheet
        if "conversion_rates" in metric_data:
            ws_conversion = wb.create_sheet("Conversion Rates")
            ws_conversion.append(["From Stage", "To Stage", "From Count", "To Count", "Conversion Rate (%)"])

            conversion_data = metric_data["conversion_rates"]
            for conversion in conversion_data:
                ws_conversion.append([
                    conversion["from_stage"].upper(),
                    conversion["to_stage"].upper(),
                    conversion["from_count"],
                    conversion["to_count"],
                    round(conversion["conversion_rate"], 2),
                ])

            # Style header
            for cell in ws_conversion[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Adjust column widths
            ws_conversion.column_dimensions["A"].width = 20
            ws_conversion.column_dimensions["B"].width = 20
            ws_conversion.column_dimensions["C"].width = 15
            ws_conversion.column_dimensions["D"].width = 15
            ws_conversion.column_dimensions["E"].width = 22

        # Skills distribution sheet
        if "skills_distribution" in metric_data:
            ws_skills = wb.create_sheet("Skills Distribution")
            ws_skills.append(["Skill", "Count", "Percentage (%)"])

            skills_data = metric_data["skills_distribution"]
            for skill, values in skills_data.items():
                ws_skills.append([
                    skill,
                    values["count"],
                    values["percentage"],
                ])

            # Style header
            for cell in ws_skills[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Adjust column widths
            ws_skills.column_dimensions["A"].width = 30
            ws_skills.column_dimensions["B"].width = 15
            ws_skills.column_dimensions["C"].width = 18

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()

        # Generate filename
        filename = self._generate_filename(
            self.REPORT_TYPE_CUSTOM_ANALYTICS,
            self.FORMAT_EXCEL
        )

        return ReportGenerationResult(
            success=True,
            report_bytes=excel_bytes,
            filename=filename,
            content_type=self.CONTENT_TYPES[self.FORMAT_EXCEL],
            file_size=len(excel_bytes),
            metadata={
                "total_candidates": data["total_candidates"],
                "metrics_count": len(data["metrics"]),
                "organization_id": data.get("organization_id"),
            },
        )

    async def _generate_custom_analytics_csv(
        self, config: ReportConfig, data: Dict[str, Any]
    ) -> ReportGenerationResult:
        """
        Generate custom analytics report in CSV format.

        Args:
            config: Report configuration
            data: Collected custom analytics data

        Returns:
            ReportGenerationResult with CSV bytes
        """
        import csv

        # Create CSV buffer
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        # Write header
        writer.writerow(["Custom Analytics Report"])
        writer.writerow([])

        # Write summary section
        writer.writerow(["Summary"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Candidates", data["total_candidates"]])
        writer.writerow(["Metrics Analyzed", len(data["metrics"])])
        writer.writerow(["Report Generated", data["summary_statistics"]["report_generated_at"]])

        if data.get("date_range"):
            writer.writerow(["Date Range Start", data["date_range"].get("start", "N/A")])
            writer.writerow(["Date Range End", data["date_range"].get("end", "N/A")])

        if data.get("organization_id"):
            writer.writerow(["Organization ID", data["organization_id"]])

        writer.writerow([])
        writer.writerow([])

        metric_data = data.get("metric_data", {})

        # Candidates by stage
        if "candidates_by_stage" in metric_data:
            writer.writerow(["Candidates by Hiring Stage"])
            writer.writerow(["Stage", "Count", "Percentage (%)"])

            stage_data = metric_data["candidates_by_stage"]
            for stage, values in sorted(stage_data.items()):
                writer.writerow([
                    stage.upper(),
                    values["count"],
                    values["percentage"],
                ])

            writer.writerow([])
            writer.writerow([])

        # Candidates by status
        if "candidates_by_status" in metric_data:
            writer.writerow(["Candidates by Status"])
            writer.writerow(["Status", "Count", "Percentage (%)"])

            status_data = metric_data["candidates_by_status"]
            for status, values in sorted(status_data.items()):
                writer.writerow([
                    status.upper(),
                    values["count"],
                    values["percentage"],
                ])

            writer.writerow([])
            writer.writerow([])

        # Time to hire
        if "time_to_hire" in metric_data:
            writer.writerow(["Time to Hire Metrics"])
            writer.writerow(["Metric", "Value"])

            time_data = metric_data["time_to_hire"]
            writer.writerow(["Average Time", time_data["avg_time_formatted"]])
            writer.writerow(["Average Time (Days)", round(time_data["avg_time_days"], 2)])
            writer.writerow(["Minimum Time", time_data["min_time_formatted"]])
            writer.writerow(["Minimum Time (Days)", round(time_data["min_time_days"], 2)])
            writer.writerow(["Maximum Time", time_data["max_time_formatted"]])
            writer.writerow(["Maximum Time (Days)", round(time_data["max_time_days"], 2)])

            writer.writerow([])
            writer.writerow([])

        # Conversion rates
        if "conversion_rates" in metric_data:
            writer.writerow(["Stage Conversion Rates"])
            writer.writerow(["From Stage", "To Stage", "From Count", "To Count", "Conversion Rate (%)"])

            conversion_data = metric_data["conversion_rates"]
            for conversion in conversion_data:
                writer.writerow([
                    conversion["from_stage"].upper(),
                    conversion["to_stage"].upper(),
                    conversion["from_count"],
                    conversion["to_count"],
                    round(conversion["conversion_rate"], 2),
                ])

            writer.writerow([])
            writer.writerow([])

        # Skills distribution
        if "skills_distribution" in metric_data:
            writer.writerow(["Top Skills Distribution"])
            writer.writerow(["Skill", "Count", "Percentage (%)"])

            skills_data = metric_data["skills_distribution"]
            for skill, values in skills_data.items():
                writer.writerow([
                    skill,
                    values["count"],
                    values["percentage"],
                ])

            writer.writerow([])
            writer.writerow([])

        # Get CSV content
        csv_content = buffer.getvalue()
        buffer.close()

        # Convert to bytes
        csv_bytes = csv_content.encode("utf-8")

        # Generate filename
        filename = self._generate_filename(
            self.REPORT_TYPE_CUSTOM_ANALYTICS,
            self.FORMAT_CSV
        )

        return ReportGenerationResult(
            success=True,
            report_bytes=csv_bytes,
            filename=filename,
            content_type=self.CONTENT_TYPES[self.FORMAT_CSV],
            file_size=len(csv_bytes),
            metadata={
                "total_candidates": data["total_candidates"],
                "metrics_count": len(data["metrics"]),
                "organization_id": data.get("organization_id"),
            },
        )

    async def _collect_eeoc_compliance_data(
        self, config: ReportConfig
    ) -> Dict[str, Any]:
        """
        Collect data for EEOC compliance report from database.

        This method aggregates demographic data to calculate:
        - Diversity metrics by gender, age group, ethnicity
        - Hiring statistics by protected class
        - Sample sizes and percentages
        - Confidence indicators for inferred data

        Args:
            config: Report configuration with filters and date_range

        Returns:
            Dictionary with aggregated EEOC compliance data

        Raises:
            ValueError: If required configuration is missing
        """
        from backend.models.demographic_inference import DemographicInference
        from backend.models.resume import Resume
        from backend.models.hiring_stage import HiringStage
        from sqlalchemy import func, and_, or_

        # Initialize data structure
        eeoc_data: Dict[str, Any] = {
            "total_candidates": 0,
            "gender_breakdown": {},
            "age_breakdown": {},
            "ethnicity_breakdown": {},
            "geographic_breakdown": {},
            "career_stage_breakdown": {},
            "education_level_breakdown": {},
            "hiring_statistics": {},
            "confidence_metrics": {},
            "date_range": config.date_range,
            "organization_id": config.organization_id,
        }

        # Build base query for resumes with demographic inferences
        query = select(
            Resume.id,
            Resume.created_at,
            DemographicInference.inferred_gender,
            DemographicInference.gender_confidence,
            DemographicInference.inferred_age_group,
            DemographicInference.age_confidence,
            DemographicInference.inferred_ethnicity,
            DemographicInference.ethnicity_confidence,
            DemographicInference.inferred_geographic_region,
            DemographicInference.geographic_confidence,
            DemographicInference.inferred_career_stage,
            DemographicInference.career_stage_confidence,
            DemographicInference.inferred_education_level,
            DemographicInference.education_confidence,
        ).select_from(Resume).outerjoin(
            DemographicInference,
            Resume.id == DemographicInference.resume_id
        )

        # Apply filters
        conditions = []

        # Organization filter
        if config.organization_id:
            conditions.append(Resume.organization_id == config.organization_id)

        # Date range filter
        if config.date_range:
            if config.date_range.get("start"):
                conditions.append(Resume.created_at >= config.date_range["start"])
            if config.date_range.get("end"):
                conditions.append(Resume.created_at <= config.date_range["end"])

        # Additional filters from config
        if config.filters:
            # Add custom filters here if needed
            pass

        if conditions:
            query = query.where(and_(*conditions))

        # Execute query
        result = await self.db.execute(query)
        rows = result.all()

        # Process results
        eeoc_data["total_candidates"] = len(rows)

        # Initialize breakdown dictionaries
        gender_counts = {}
        age_counts = {}
        ethnicity_counts = {}
        geographic_counts = {}
        career_stage_counts = {}
        education_level_counts = {}

        # Track confidence scores
        gender_confidences = []
        age_confidences = []
        ethnicity_confidences = []

        for row in rows:
            # Gender breakdown
            gender = row.inferred_gender or "Unknown"
            gender_counts[gender] = gender_counts.get(gender, 0) + 1
            if row.gender_confidence:
                gender_confidences.append(row.gender_confidence)

            # Age breakdown
            age_group = row.inferred_age_group or "Unknown"
            age_counts[age_group] = age_counts.get(age_group, 0) + 1
            if row.age_confidence:
                age_confidences.append(row.age_confidence)

            # Ethnicity breakdown
            ethnicity = row.inferred_ethnicity or "Unknown"
            ethnicity_counts[ethnicity] = ethnicity_counts.get(ethnicity, 0) + 1
            if row.ethnicity_confidence:
                ethnicity_confidences.append(row.ethnicity_confidence)

            # Geographic breakdown
            region = row.inferred_geographic_region or "Unknown"
            geographic_counts[region] = geographic_counts.get(region, 0) + 1

            # Career stage breakdown
            career_stage = row.inferred_career_stage or "Unknown"
            career_stage_counts[career_stage] = career_stage_counts.get(career_stage, 0) + 1

            # Education level breakdown
            education_level = row.inferred_education_level or "Unknown"
            education_level_counts[education_level] = education_level_counts.get(education_level, 0) + 1

        # Calculate percentages and format breakdowns
        total = eeoc_data["total_candidates"]

        # Gender breakdown with percentages
        eeoc_data["gender_breakdown"] = {
            gender: {
                "count": count,
                "percentage": round((count / total * 100), 2) if total > 0 else 0,
            }
            for gender, count in sorted(gender_counts.items())
        }

        # Age breakdown with percentages
        eeoc_data["age_breakdown"] = {
            age_group: {
                "count": count,
                "percentage": round((count / total * 100), 2) if total > 0 else 0,
            }
            for age_group, count in sorted(age_counts.items())
        }

        # Ethnicity breakdown with percentages
        eeoc_data["ethnicity_breakdown"] = {
            ethnicity: {
                "count": count,
                "percentage": round((count / total * 100), 2) if total > 0 else 0,
            }
            for ethnicity, count in sorted(ethnicity_counts.items())
        }

        # Geographic breakdown with percentages
        eeoc_data["geographic_breakdown"] = {
            region: {
                "count": count,
                "percentage": round((count / total * 100), 2) if total > 0 else 0,
            }
            for region, count in sorted(geographic_counts.items())
        }

        # Career stage breakdown with percentages
        eeoc_data["career_stage_breakdown"] = {
            stage: {
                "count": count,
                "percentage": round((count / total * 100), 2) if total > 0 else 0,
            }
            for stage, count in sorted(career_stage_counts.items())
        }

        # Education level breakdown with percentages
        eeoc_data["education_level_breakdown"] = {
            level: {
                "count": count,
                "percentage": round((count / total * 100), 2) if total > 0 else 0,
            }
            for level, count in sorted(education_level_counts.items())
        }

        # Calculate confidence metrics
        eeoc_data["confidence_metrics"] = {
            "gender_avg_confidence": round(sum(gender_confidences) / len(gender_confidences), 3) if gender_confidences else 0.0,
            "age_avg_confidence": round(sum(age_confidences) / len(age_confidences), 3) if age_confidences else 0.0,
            "ethnicity_avg_confidence": round(sum(ethnicity_confidences) / len(ethnicity_confidences), 3) if ethnicity_confidences else 0.0,
            "gender_inferred_count": len(gender_confidences),
            "age_inferred_count": len(age_confidences),
            "ethnicity_inferred_count": len(ethnicity_confidences),
        }

        # Get hiring statistics (hired candidates)
        hired_query = select(func.count(HiringStage.id)).where(
            HiringStage.stage_name == "hired"
        )
        if conditions:
            # Join with Resume to apply filters
            hired_query = hired_query.select_from(HiringStage).join(
                Resume, HiringStage.resume_id == Resume.id
            ).where(and_(*conditions))

        hired_result = await self.db.execute(hired_query)
        hired_count = hired_result.scalar() or 0

        eeoc_data["hiring_statistics"] = {
            "total_hired": hired_count,
            "hiring_rate": round((hired_count / total * 100), 2) if total > 0 else 0,
        }

        logger.info(
            f"Collected EEOC compliance data: {total} candidates, "
            f"{hired_count} hired, "
            f"{len(gender_counts)} gender categories, "
            f"{len(ethnicity_counts)} ethnicity categories"
        )

        return eeoc_data

    async def _generate_eeoc_compliance_pdf(
        self, config: ReportConfig, data: Dict[str, Any]
    ) -> ReportGenerationResult:
        """
        Generate EEOC compliance report in PDF format.

        Args:
            config: Report configuration
            data: Collected EEOC compliance data

        Returns:
            ReportGenerationResult with PDF bytes
        """
        if not self._reportlab_available:
            return ReportGenerationResult(
                success=False,
                error_message="PDF generation requires reportlab library",
            )

        from reportlab.lib.pagesizes import A4, letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.pdfgen import canvas

        # Create PDF buffer
        buffer = io.BytesIO()

        # Determine page size
        pagesize = letter if config.page_format == self.PAGE_FORMAT_LETTER else A4

        # Create document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=pagesize,
            rightMargin=0.75 * inch,
            leftMargin=0.75 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
        )

        # Get styles
        styles = getSampleStyleSheet()
        story = []

        # Create custom styles
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#2c3e50"),
            spaceAfter=30,
            alignment=TA_CENTER,
        )

        heading_style = ParagraphStyle(
            "CustomHeading",
            parent=styles["Heading2"],
            fontSize=16,
            textColor=colors.HexColor("#34495e"),
            spaceAfter=12,
            spaceBefore=12,
        )

        normal_style = ParagraphStyle(
            "CustomNormal",
            parent=styles["Normal"],
            fontSize=11,
            textColor=colors.HexColor("#2c3e50"),
            spaceAfter=12,
        )

        # Title
        title = config.title or "EEOC Compliance Report"
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 0.2 * inch))

        # Summary Section
        story.append(Paragraph("Executive Summary", heading_style))
        summary_data = [
            ["Total Candidates", str(data["total_candidates"])],
            ["Total Hired", str(data["hiring_statistics"]["total_hired"])],
            ["Hiring Rate", f"{data['hiring_statistics']['hiring_rate']}%"],
        ]

        if data.get("date_range"):
            if data["date_range"].get("start"):
                summary_data.append(["Date Range Start", data["date_range"]["start"]])
            if data["date_range"].get("end"):
                summary_data.append(["Date Range End", data["date_range"]["end"]])

        summary_table = Table(summary_data, colWidths=[3 * inch, 3 * inch])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#34495e")),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 11),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
            ("BACKGROUND", (1, 0), (1, -1), colors.HexColor("#ecf0f1")),
            ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#bdc3c7")),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3 * inch))

        # Gender Diversity Section
        story.append(Paragraph("Gender Diversity Metrics", heading_style))
        story.append(Paragraph(
            f"Average Confidence Score: {data['confidence_metrics']['gender_avg_confidence']:.1%}",
            normal_style
        ))

        gender_data = [["Gender", "Count", "Percentage"]]
        for gender, stats in data["gender_breakdown"].items():
            gender_data.append([
                gender.replace("_", " ").title(),
                str(stats["count"]),
                f"{stats['percentage']}%",
            ])

        gender_table = Table(gender_data, colWidths=[2.5 * inch, 1.5 * inch, 1.5 * inch])
        gender_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#ecf0f1")),
            ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#bdc3c7")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
        ]))
        story.append(gender_table)
        story.append(Spacer(1, 0.3 * inch))

        # Age Distribution Section
        story.append(Paragraph("Age Distribution", heading_style))
        story.append(Paragraph(
            f"Average Confidence Score: {data['confidence_metrics']['age_avg_confidence']:.1%}",
            normal_style
        ))

        age_data = [["Age Group", "Count", "Percentage"]]
        for age_group, stats in data["age_breakdown"].items():
            age_data.append([
                age_group.replace("_", " ").title(),
                str(stats["count"]),
                f"{stats['percentage']}%",
            ])

        age_table = Table(age_data, colWidths=[2.5 * inch, 1.5 * inch, 1.5 * inch])
        age_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#ecf0f1")),
            ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#bdc3c7")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
        ]))
        story.append(age_table)
        story.append(Spacer(1, 0.3 * inch))

        # Ethnicity Diversity Section
        story.append(Paragraph("Ethnicity Diversity Metrics", heading_style))
        story.append(Paragraph(
            f"Average Confidence Score: {data['confidence_metrics']['ethnicity_avg_confidence']:.1%}",
            normal_style
        ))

        ethnicity_data = [["Ethnicity", "Count", "Percentage"]]
        for ethnicity, stats in data["ethnicity_breakdown"].items():
            ethnicity_data.append([
                ethnicity.replace("_", " ").title(),
                str(stats["count"]),
                f"{stats['percentage']}%",
            ])

        ethnicity_table = Table(ethnicity_data, colWidths=[2.5 * inch, 1.5 * inch, 1.5 * inch])
        ethnicity_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#ecf0f1")),
            ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#bdc3c7")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
        ]))
        story.append(ethnicity_table)
        story.append(Spacer(1, 0.3 * inch))

        # Geographic Distribution Section (if data available)
        if data["geographic_breakdown"]:
            story.append(Paragraph("Geographic Distribution", heading_style))

            geographic_data = [["Region", "Count", "Percentage"]]
            for region, stats in data["geographic_breakdown"].items():
                geographic_data.append([
                    region.replace("_", " ").title(),
                    str(stats["count"]),
                    f"{stats['percentage']}%",
                ])

            geographic_table = Table(geographic_data, colWidths=[2.5 * inch, 1.5 * inch, 1.5 * inch])
            geographic_table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#34495e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#ecf0f1")),
                ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#bdc3c7")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
            ]))
            story.append(geographic_table)
            story.append(Spacer(1, 0.3 * inch))

        # Footer with generation timestamp
        footer_style = ParagraphStyle(
            "Footer",
            parent=normal_style,
            fontSize=9,
            textColor=colors.HexColor("#7f8c8d"),
            alignment=TA_CENTER,
        )
        timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
        footer_text = f"Generated: {timestamp}"
        if data.get("organization_id"):
            footer_text += f" | Organization: {data['organization_id']}"
        footer_text += "<br/>Note: Demographic data is AI-inferred and used for compliance monitoring only."
        story.append(Spacer(1, 0.5 * inch))
        story.append(Paragraph(footer_text, footer_style))

        # Build PDF
        doc.build(story)

        # Get PDF bytes
        pdf_bytes = buffer.getvalue()
        buffer.close()

        # Generate filename
        filename = self._generate_filename(
            self.REPORT_TYPE_EEOC_COMPLIANCE,
            self.FORMAT_PDF
        )

        return ReportGenerationResult(
            success=True,
            report_bytes=pdf_bytes,
            filename=filename,
            content_type=self.CONTENT_TYPES[self.FORMAT_PDF],
            file_size=len(pdf_bytes),
            metadata={
                "total_candidates": data["total_candidates"],
                "total_hired": data["hiring_statistics"]["total_hired"],
                "organization_id": data.get("organization_id"),
                "date_range": data.get("date_range"),
            },
        )

    async def _generate_eeoc_compliance_excel(
        self, config: ReportConfig, data: Dict[str, Any]
    ) -> ReportGenerationResult:
        """
        Generate EEOC compliance report in Excel format.

        Args:
            config: Report configuration
            data: Collected EEOC compliance data

        Returns:
            ReportGenerationResult with Excel bytes
        """
        if not self._openpyxl_available:
            return ReportGenerationResult(
                success=False,
                error_message="Excel generation requires openpyxl library",
            )

        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["EEOC Compliance Report"])
        ws_summary.append([])
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Total Candidates", data["total_candidates"]])
        ws_summary.append(["Total Hired", data["hiring_statistics"]["total_hired"]])
        ws_summary.append(["Hiring Rate", f"{data['hiring_statistics']['hiring_rate']}%"])

        if data.get("date_range"):
            if data["date_range"].get("start"):
                ws_summary.append(["Date Range Start", data["date_range"]["start"]])
            if data["date_range"].get("end"):
                ws_summary.append(["Date Range End", data["date_range"]["end"]])

        # Style summary sheet
        ws_summary["A1"].font = Font(size=16, bold=True)
        ws_summary["A3"].font = Font(bold=True)
        ws_summary["B3"].font = Font(bold=True)
        ws_summary["A3"].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        ws_summary["B3"].fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        ws_summary["A3"].font = Font(color="FFFFFF", bold=True)
        ws_summary["B3"].font = Font(color="FFFFFF", bold=True)

        # Adjust column widths
        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 40

        # Create Gender Diversity sheet
        ws_gender = wb.create_sheet("Gender Diversity")
        ws_gender.append(["Gender", "Count", "Percentage"])

        for gender, stats in data["gender_breakdown"].items():
            ws_gender.append([
                gender.replace("_", " ").title(),
                stats["count"],
                f"{stats['percentage']}%",
            ])

        # Add confidence metric
        ws_gender.append([])
        ws_gender.append(["Average Confidence Score", f"{data['confidence_metrics']['gender_avg_confidence']:.1%}"])

        # Style gender sheet
        for cell in ws_gender[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        ws_gender.column_dimensions["A"].width = 20
        ws_gender.column_dimensions["B"].width = 15
        ws_gender.column_dimensions["C"].width = 15

        # Create Age Distribution sheet
        ws_age = wb.create_sheet("Age Distribution")
        ws_age.append(["Age Group", "Count", "Percentage"])

        for age_group, stats in data["age_breakdown"].items():
            ws_age.append([
                age_group.replace("_", " ").title(),
                stats["count"],
                f"{stats['percentage']}%",
            ])

        # Add confidence metric
        ws_age.append([])
        ws_age.append(["Average Confidence Score", f"{data['confidence_metrics']['age_avg_confidence']:.1%}"])

        # Style age sheet
        for cell in ws_age[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        ws_age.column_dimensions["A"].width = 20
        ws_age.column_dimensions["B"].width = 15
        ws_age.column_dimensions["C"].width = 15

        # Create Ethnicity Diversity sheet
        ws_ethnicity = wb.create_sheet("Ethnicity Diversity")
        ws_ethnicity.append(["Ethnicity", "Count", "Percentage"])

        for ethnicity, stats in data["ethnicity_breakdown"].items():
            ws_ethnicity.append([
                ethnicity.replace("_", " ").title(),
                stats["count"],
                f"{stats['percentage']}%",
            ])

        # Add confidence metric
        ws_ethnicity.append([])
        ws_ethnicity.append(["Average Confidence Score", f"{data['confidence_metrics']['ethnicity_avg_confidence']:.1%}"])

        # Style ethnicity sheet
        for cell in ws_ethnicity[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")

        ws_ethnicity.column_dimensions["A"].width = 25
        ws_ethnicity.column_dimensions["B"].width = 15
        ws_ethnicity.column_dimensions["C"].width = 15

        # Create Geographic Distribution sheet (if data available)
        if data["geographic_breakdown"]:
            ws_geographic = wb.create_sheet("Geographic Distribution")
            ws_geographic.append(["Region", "Count", "Percentage"])

            for region, stats in data["geographic_breakdown"].items():
                ws_geographic.append([
                    region.replace("_", " ").title(),
                    stats["count"],
                    f"{stats['percentage']}%",
                ])

            # Style geographic sheet
            for cell in ws_geographic[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

            ws_geographic.column_dimensions["A"].width = 25
            ws_geographic.column_dimensions["B"].width = 15
            ws_geographic.column_dimensions["C"].width = 15

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()

        # Generate filename
        filename = self._generate_filename(
            self.REPORT_TYPE_EEOC_COMPLIANCE,
            self.FORMAT_EXCEL
        )

        return ReportGenerationResult(
            success=True,
            report_bytes=excel_bytes,
            filename=filename,
            content_type=self.CONTENT_TYPES[self.FORMAT_EXCEL],
            file_size=len(excel_bytes),
            metadata={
                "total_candidates": data["total_candidates"],
                "total_hired": data["hiring_statistics"]["total_hired"],
                "organization_id": data.get("organization_id"),
                "date_range": data.get("date_range"),
            },
        )

    async def _generate_eeoc_compliance_csv(
        self, config: ReportConfig, data: Dict[str, Any]
    ) -> ReportGenerationResult:
        """
        Generate EEOC compliance report in CSV format.

        Args:
            config: Report configuration
            data: Collected EEOC compliance data

        Returns:
            ReportGenerationResult with CSV bytes
        """
        import csv

        # Create CSV buffer
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        # Write title and summary
        writer.writerow(["EEOC Compliance Report"])
        writer.writerow([])
        writer.writerow(["Summary Metrics"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Candidates", data["total_candidates"]])
        writer.writerow(["Total Hired", data["hiring_statistics"]["total_hired"]])
        writer.writerow(["Hiring Rate", f"{data['hiring_statistics']['hiring_rate']}%"])

        if data.get("date_range"):
            if data["date_range"].get("start"):
                writer.writerow(["Date Range Start", data["date_range"]["start"]])
            if data["date_range"].get("end"):
                writer.writerow(["Date Range End", data["date_range"]["end"]])

        writer.writerow([])
        writer.writerow([])

        # Gender Diversity section
        writer.writerow(["Gender Diversity Metrics"])
        writer.writerow(["Average Confidence Score", f"{data['confidence_metrics']['gender_avg_confidence']:.1%}"])
        writer.writerow(["Gender", "Count", "Percentage"])

        for gender, stats in data["gender_breakdown"].items():
            writer.writerow([
                gender.replace("_", " ").title(),
                stats["count"],
                f"{stats['percentage']}%",
            ])

        writer.writerow([])
        writer.writerow([])

        # Age Distribution section
        writer.writerow(["Age Distribution"])
        writer.writerow(["Average Confidence Score", f"{data['confidence_metrics']['age_avg_confidence']:.1%}"])
        writer.writerow(["Age Group", "Count", "Percentage"])

        for age_group, stats in data["age_breakdown"].items():
            writer.writerow([
                age_group.replace("_", " ").title(),
                stats["count"],
                f"{stats['percentage']}%",
            ])

        writer.writerow([])
        writer.writerow([])

        # Ethnicity Diversity section
        writer.writerow(["Ethnicity Diversity Metrics"])
        writer.writerow(["Average Confidence Score", f"{data['confidence_metrics']['ethnicity_avg_confidence']:.1%}"])
        writer.writerow(["Ethnicity", "Count", "Percentage"])

        for ethnicity, stats in data["ethnicity_breakdown"].items():
            writer.writerow([
                ethnicity.replace("_", " ").title(),
                stats["count"],
                f"{stats['percentage']}%",
            ])

        writer.writerow([])
        writer.writerow([])

        # Geographic Distribution section (if data available)
        if data["geographic_breakdown"]:
            writer.writerow(["Geographic Distribution"])
            writer.writerow(["Region", "Count", "Percentage"])

            for region, stats in data["geographic_breakdown"].items():
                writer.writerow([
                    region.replace("_", " ").title(),
                    stats["count"],
                    f"{stats['percentage']}%",
                ])

            writer.writerow([])
            writer.writerow([])

        # Footer
        writer.writerow(["Note: Demographic data is AI-inferred and used for compliance monitoring only."])
        writer.writerow(["Generated:", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")])

        # Get CSV content
        csv_content = buffer.getvalue()
        buffer.close()

        # Convert to bytes
        csv_bytes = csv_content.encode("utf-8")

        # Generate filename
        filename = self._generate_filename(
            self.REPORT_TYPE_EEOC_COMPLIANCE,
            self.FORMAT_CSV
        )

        return ReportGenerationResult(
            success=True,
            report_bytes=csv_bytes,
            filename=filename,
            content_type=self.CONTENT_TYPES[self.FORMAT_CSV],
            file_size=len(csv_bytes),
            metadata={
                "total_candidates": data["total_candidates"],
                "total_hired": data["hiring_statistics"]["total_hired"],
                "organization_id": data.get("organization_id"),
                "date_range": data.get("date_range"),
            },
        )

    def _generate_filename(
        self,
        report_type: str,
        output_format: str,
        timestamp: Optional[str] = None,
    ) -> str:
        """
        Generate filename for report.

        Args:
            report_type: Type of report
            output_format: Output format
            timestamp: Optional timestamp (default: current time)

        Returns:
            Suggested filename for the report

        Example:
            >>> service._generate_filename("candidate_summary", "pdf")
            "candidate_summary_2024-03-21_143045.pdf"
        """
        if timestamp is None:
            timestamp = datetime.utcnow().strftime("%Y-%m-%d_%H%M%S")

        extension_map = {
            self.FORMAT_PDF: "pdf",
            self.FORMAT_EXCEL: "xlsx",
            self.FORMAT_CSV: "csv",
        }

        extension = extension_map.get(output_format, "bin")
        return f"{report_type}_{timestamp}.{extension}"


# Singleton pattern for global instance
def get_report_template_service(db: AsyncSession) -> ReportTemplateService:
    """
    Get or create the report template service instance.

    Args:
        db: Database session

    Returns:
        ReportTemplateService instance

    Example:
        >>> from sqlalchemy.ext.asyncio import AsyncSession
        >>> service = get_report_template_service(db)
        >>> result = await service.generate_report(config, data)
    """
    return ReportTemplateService(db)
