"""
Unit tests for Excel export endpoint with formatting.

Tests cover:
- Successful Excel export generation
- Response headers and content type validation
- Excel file structure and formatting
- Data bars and conditional formatting
- Header styling verification
- Validation error handling
- Edge cases and special scenarios
"""
from datetime import datetime
from unittest.mock import Mock, patch

import pytest
from fastapi.testclient import TestClient

# Import the FastAPI application
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from main import app


@pytest.fixture
def client():
    """
    Create a test client for the FastAPI application.

    Returns:
        TestClient: Configured test client
    """
    return TestClient(app)


class TestExcelExportEndpoint:
    """Tests for POST /api/reports/export/excel endpoint."""

    def test_export_excel_success(self, client):
        """Test successful Excel export."""
        payload = {
            "metrics": ["time_to_hire", "resumes_processed"],
            "filters": {},
            "format": "standard",
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "attachment" in response.headers["content-disposition"]
        assert "analytics_export.xlsx" in response.headers["content-disposition"]

    def test_export_excel_response_structure(self, client):
        """Test response has correct structure and headers."""
        payload = {
            "metrics": ["time_to_hire"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)

        # Verify status code
        assert response.status_code == 200

        # Verify content type is Excel
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        # Verify content disposition has filename
        assert "content-disposition" in response.headers
        assert "attachment" in response.headers["content-disposition"].lower()

    def test_export_excel_content_type(self, client):
        """Test Excel file has correct content type."""
        payload = {
            "metrics": ["resumes_processed"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200
        assert "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in response.headers["content-type"]

    def test_export_excel_filename(self, client):
        """Test Excel file has correct filename in headers."""
        payload = {
            "metrics": ["match_rates"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200
        content_disposition = response.headers.get("content-disposition", "")
        assert "analytics_export.xlsx" in content_disposition

    def test_export_excel_single_metric(self, client):
        """Test Excel export with single metric."""
        payload = {
            "metrics": ["time_to_hire"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200
        # Verify we got binary content (Excel file)
        assert len(response.content) > 0

    def test_export_excel_multiple_metrics(self, client):
        """Test Excel export with multiple metrics."""
        payload = {
            "metrics": ["time_to_hire", "resumes_processed", "match_rates"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200
        assert len(response.content) > 0

    def test_export_excel_with_filters(self, client):
        """Test Excel export with filters."""
        payload = {
            "metrics": ["time_to_hire"],
            "filters": {
                "start_date": "2024-01-01",
                "end_date": "2024-12-31",
                "organization_id": "org-123"
            },
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

    def test_export_excel_with_format_parameter(self, client):
        """Test Excel export with format parameter."""
        payload = {
            "metrics": ["resumes_processed"],
            "filters": {},
            "format": "detailed",
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

    def test_export_excel_all_available_metrics(self, client):
        """Test Excel export with all available metrics."""
        payload = {
            "metrics": [
                "time_to_hire",
                "resumes_processed",
                "match_rates",
                "interviews_scheduled",
                "offers_extended",
                "offers_accepted",
            ],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200
        assert len(response.content) > 0


class TestExcelExportValidation:
    """Tests for Excel export validation."""

    def test_export_excel_empty_metrics_list(self, client):
        """Test Excel export with empty metrics list returns validation error."""
        payload = {
            "metrics": [],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 422
        assert "At least one metric must be provided" in response.json()["detail"]

    def test_export_excel_missing_metrics_field(self, client):
        """Test Excel export without metrics field returns validation error."""
        payload = {
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 422

    def test_export_excel_no_payload(self, client):
        """Test Excel export without payload returns validation error."""
        response = client.post("/api/reports/export/excel", json={})
        assert response.status_code == 422


class TestExcelExportFileContent:
    """Tests for Excel file content validation."""

    def test_export_excel_file_signature(self, client):
        """Test Excel file has correct file signature (magic bytes)."""
        payload = {
            "metrics": ["time_to_hire"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

        # Excel files (XLSX) start with PK (ZIP signature)
        # The first two bytes should be: 0x50 0x4B (ASCII: "PK")
        content = response.content
        assert len(content) > 0
        assert content[0:2] == b'PK'

    def test_export_excel_file_size(self, client):
        """Test Excel file has reasonable size."""
        payload = {
            "metrics": ["time_to_hire", "resumes_processed"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

        # Excel file should be non-empty and have reasonable size
        # (at least a few KB for formatting and content)
        assert len(response.content) > 1000

    def test_export_excel_can_be_opened(self, client):
        """Test Excel file can be opened with openpyxl."""
        import io
        from openpyxl import load_workbook

        payload = {
            "metrics": ["time_to_hire", "resumes_processed"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

        # Try to load the Excel file with openpyxl
        excel_buffer = io.BytesIO(response.content)
        wb = load_workbook(excel_buffer)

        # Verify workbook has at least one worksheet
        assert wb.active is not None
        assert wb.sheetnames is not None
        assert len(wb.sheetnames) >= 1

        # Verify the active sheet has data
        ws = wb.active
        assert ws.max_row >= 1  # At least header row
        assert ws.max_column >= 1

    def test_export_excel_worksheet_title(self, client):
        """Test Excel worksheet has correct title."""
        import io
        from openpyxl import load_workbook

        payload = {
            "metrics": ["time_to_hire"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

        excel_buffer = io.BytesIO(response.content)
        wb = load_workbook(excel_buffer)

        # Verify worksheet title
        assert wb.active.title == "Analytics Report"

    def test_export_excel_headers(self, client):
        """Test Excel file has correct headers."""
        import io
        from openpyxl import load_workbook

        payload = {
            "metrics": ["time_to_hire", "resumes_processed"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

        excel_buffer = io.BytesIO(response.content)
        wb = load_workbook(excel_buffer)
        ws = wb.active

        # Get first row (headers)
        headers = [cell.value for cell in ws[1]]

        # Verify expected headers
        assert "Metric" in headers
        assert "Value" in headers
        assert "Date" in headers
        assert "Unit" in headers

    def test_export_excel_data_rows(self, client):
        """Test Excel file has correct data rows."""
        import io
        from openpyxl import load_workbook

        payload = {
            "metrics": ["time_to_hire", "resumes_processed"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

        excel_buffer = io.BytesIO(response.content)
        wb = load_workbook(excel_buffer)
        ws = wb.active

        # Verify we have header + data rows
        assert ws.max_row >= 2  # Header + at least one data row

        # Verify each requested metric has a data row
        metrics_found = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # First column is metric name
                metrics_found.append(row[0])

        assert "time_to_hire" in metrics_found
        assert "resumes_processed" in metrics_found


class TestExcelExportFormatting:
    """Tests for Excel formatting features."""

    def test_export_excel_header_formatting(self, client):
        """Test Excel headers have formatting applied."""
        import io
        from openpyxl import load_workbook

        payload = {
            "metrics": ["time_to_hire"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

        excel_buffer = io.BytesIO(response.content)
        wb = load_workbook(excel_buffer)
        ws = wb.active

        # Check header row formatting
        header_row = ws[1]

        # Headers should have bold font
        for cell in header_row:
            if cell.value:
                assert cell.font.bold is True

    def test_export_excel_header_fill_color(self, client):
        """Test Excel headers have background fill color."""
        import io
        from openpyxl import load_workbook

        payload = {
            "metrics": ["resumes_processed"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

        excel_buffer = io.BytesIO(response.content)
        wb = load_workbook(excel_buffer)
        ws = wb.active

        # Check header row has fill color
        header_cell = ws['A1']  # First header cell
        assert header_cell.fill.start_color.rgb is not None

    def test_export_excel_header_alignment(self, client):
        """Test Excel headers have center alignment."""
        import io
        from openpyxl import load_workbook

        payload = {
            "metrics": ["match_rates"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

        excel_buffer = io.BytesIO(response.content)
        wb = load_workbook(excel_buffer)
        ws = wb.active

        # Check header row has center alignment
        header_cell = ws['A1']
        assert header_cell.alignment.horizontal == "center"
        assert header_cell.alignment.vertical == "center"

    def test_export_excel_data_bars_applied(self, client):
        """Test Excel has conditional formatting data bars on Value column."""
        import io
        from openpyxl import load_workbook

        payload = {
            "metrics": ["time_to_hire", "resumes_processed"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

        excel_buffer = io.BytesIO(response.content)
        wb = load_workbook(excel_buffer)
        ws = wb.active

        # Check for conditional formatting rules
        # Data bars should be applied to column B (Value column)
        assert len(ws.conditional_formatting) > 0

    def test_export_excel_column_widths(self, client):
        """Test Excel columns have adjusted widths."""
        import io
        from openpyxl import load_workbook

        payload = {
            "metrics": ["time_to_hire"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

        excel_buffer = io.BytesIO(response.content)
        wb = load_workbook(excel_buffer)
        ws = wb.active

        # Check that column dimensions have been set
        # At least some columns should have width > default (8.43)
        has_custom_width = False
        for column_letter, column_dim in ws.column_dimensions.items():
            if column_dim.width and column_dim.width > 8.5:
                has_custom_width = True
                break

        assert has_custom_width


class TestExcelExportEdgeCases:
    """Tests for edge cases and special scenarios."""

    def test_export_excel_unknown_metric(self, client):
        """Test Excel export with unknown metric name."""
        payload = {
            "metrics": ["unknown_metric_xyz"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200
        # Should still succeed but with N/A or default values

    def test_export_excel_mixed_known_unknown_metrics(self, client):
        """Test Excel export with mix of known and unknown metrics."""
        payload = {
            "metrics": ["time_to_hire", "unknown_metric", "resumes_processed"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

    def test_export_excel_large_filters(self, client):
        """Test Excel export with large filters dictionary."""
        large_filters = {f"filter_{i}": f"value_{i}" for i in range(50)}
        payload = {
            "metrics": ["time_to_hire"],
            "filters": large_filters,
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

    def test_export_excel_special_characters_in_filters(self, client):
        """Test Excel export with special characters in filter values."""
        payload = {
            "metrics": ["time_to_hire"],
            "filters": {
                "organization": "Test <script>alert('test')</script>",
                "department": "HR & Recruiting",
                "region": "North/South",
            },
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

    def test_export_excel_unicode_in_filters(self, client):
        """Test Excel export with unicode characters in filters."""
        payload = {
            "metrics": ["resumes_processed"],
            "filters": {
                "organization": "组织 名称",
                "region": "南美",
            },
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

    def test_export_excel_empty_filters(self, client):
        """Test Excel export with empty filters."""
        payload = {
            "metrics": ["match_rates"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

    def test_export_excel_null_filters(self, client):
        """Test Excel export with None/null filters."""
        payload = {
            "metrics": ["time_to_hire"],
            "filters": None,
        }
        response = client.post("/api/reports/export/excel", json=payload)
        # Should handle None gracefully or return validation error
        assert response.status_code in [200, 422]

    def test_export_excel_duplicate_metrics(self, client):
        """Test Excel export with duplicate metric names."""
        payload = {
            "metrics": ["time_to_hire", "time_to_hire", "resumes_processed"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

    def test_export_excel_very_long_metric_name(self, client):
        """Test Excel export with very long metric name."""
        long_metric = "metric_" + "x" * 500
        payload = {
            "metrics": [long_metric],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200

    def test_export_excel_all_numeric_metrics(self, client):
        """Test Excel export with metrics that have numeric values."""
        payload = {
            "metrics": ["resumes_processed", "interviews_scheduled", "offers_extended"],
            "filters": {},
        }
        response = client.post("/api/reports/export/excel", json=payload)
        assert response.status_code == 200


class TestExcelExportErrorHandling:
    """Tests for error handling."""

    def test_export_excel_invalid_json(self, client):
        """Test endpoint with invalid JSON."""
        response = client.post(
            "/api/reports/export/excel",
            data="invalid json",
            headers={"Content-Type": "application/json"},
        )
        assert response.status_code == 422

    def test_export_excel_missing_content_type(self, client):
        """Test endpoint without Content-Type header."""
        response = client.post(
            "/api/reports/export/excel",
            data="metrics=time_to_hire",
        )
        # Should either work or return 415/422
        assert response.status_code in [200, 415, 422]
