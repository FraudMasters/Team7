"""
Excel/CSV export tasks for Analytics Service.

# Русский комментарий:
Этот модуль предоставляет задачи Celery для экспорта данных аналитики в форматы Excel и CSV.
Поддерживает различные форматы данных с настраиваемыми опциями экспорта.
"""
import logging
import time
import csv
from typing import Dict, Any, Optional, List
from datetime import datetime
from io import BytesIO, StringIO

from celery import shared_task
from celery.exceptions import SoftTimeLimitExceeded

from config import get_settings

logger = logging.getLogger(__name__)
settings = get_settings()


def generate_excel_export(
    data: Dict[str, Any],
    sheet_name: str = "Analytics Data",
    include_metadata: bool = True,
) -> Optional[bytes]:
    """
    Generate Excel file from analytics data.

    This function creates an Excel workbook with analytics data,
    optionally including metadata sheet with generation details.

    Args:
        data: Analytics data dictionary containing:
            - metrics: Dictionary of metric name -> value
            - dimensions: Dictionary of dimension name -> data
            - summary: Text summary of the report
            - rows: Optional list of row data for tabular export
        sheet_name: Name of the main data sheet
        include_metadata: Whether to include metadata sheet

    Returns:
        Excel file as bytes, or None if generation fails

    Example:
        >>> data = {
        ...     "metrics": {"time_to_hire": 15.2},
        ...     "rows": [{"name": "John", "value": 100}]
        ... }
        >>> excel_bytes = generate_excel_export(data)
        >>> excel_bytes is not None
        True
    """
    try:
        # Try to import pandas and openpyxl for Excel generation
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font, PatternFill, Alignment

            logger.info("Using pandas and openpyxl for Excel generation")

            # Create Excel writer
            output = BytesIO()

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Process tabular data if available
                if "rows" in data and isinstance(data["rows"], list) and len(data["rows"]) > 0:
                    # Create DataFrame from rows
                    df = pd.DataFrame(data["rows"])
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

                    # Style the header
                    worksheet = writer.sheets[sheet_name]
                    for cell in worksheet[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")

                else:
                    # Create summary sheet from metrics and dimensions
                    summary_data = []

                    # Add metrics
                    if "metrics" in data:
                        summary_data.append(["Metric", "Value"])
                        for metric_name, metric_value in data["metrics"].items():
                            if isinstance(metric_value, dict):
                                # Handle nested metrics
                                for k, v in metric_value.items():
                                    summary_data.append([f"{metric_name}.{k}", v])
                            else:
                                summary_data.append([metric_name, metric_value])

                    # Add dimensions
                    if "dimensions" in data and len(summary_data) > 0:
                        summary_data.append([])  # Empty row separator
                        summary_data.append(["Dimension", "Value"])
                        for dim_name, dim_data in data["dimensions"].items():
                            if isinstance(dim_data, dict):
                                for k, v in dim_data.items():
                                    summary_data.append([f"{dim_name}.{k}", v])
                            else:
                                summary_data.append([dim_name, dim_data])

                    if summary_data:
                        df = pd.DataFrame(summary_data)
                        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                # Add metadata sheet if requested
                if include_metadata:
                    metadata = {
                        "Field": [
                            "Generated At",
                            "Report Name",
                            "Summary",
                            "Total Metrics",
                            "Total Dimensions",
                        ],
                        "Value": [
                            data.get("generated_at", datetime.utcnow().isoformat()),
                            data.get("report_name", "Analytics Export"),
                            data.get("summary", "N/A"),
                            len(data.get("metrics", {})),
                            len(data.get("dimensions", {})),
                        ]
                    }

                    metadata_df = pd.DataFrame(metadata)
                    metadata_df.to_excel(writer, sheet_name="Metadata", index=False)

                    # Style metadata header
                    metadata_sheet = writer.sheets["Metadata"]
                    for cell in metadata_sheet[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")

            excel_bytes = output.getvalue()
            logger.info(f"Excel file generated ({len(excel_bytes)} bytes)")
            return excel_bytes

        except ImportError as ie:
            # Fallback: Create simple CSV-based Excel representation
            logger.warning(
                f"pandas/openpyxl not available ({ie}), using fallback CSV representation. "
                "Install pandas and openpyxl for proper Excel generation: "
                "pip install pandas openpyxl"
            )

            # Create CSV content as fallback
            csv_content = generate_csv_export(data)
            if csv_content:
                # Return CSV as bytes (can be opened in Excel)
                return csv_content.encode('utf-8')

            return None

    except Exception as e:
        logger.error(f"Failed to generate Excel file: {e}", exc_info=True)
        return None


def generate_csv_export(
    data: Dict[str, Any],
    delimiter: str = ",",
    include_header: bool = True,
) -> Optional[str]:
    """
    Generate CSV file from analytics data.

    This function creates a CSV representation of analytics data
    with configurable delimiter and header options.

    Args:
        data: Analytics data dictionary containing:
            - metrics: Dictionary of metric name -> value
            - dimensions: Dictionary of dimension name -> data
            - rows: Optional list of row data for tabular export
        delimiter: CSV delimiter (default: comma)
        include_header: Whether to include header row

    Returns:
        CSV content as string, or None if generation fails

    Example:
        >>> data = {
        ...     "rows": [{"name": "John", "value": 100}]
        ... }
        >>> csv_str = generate_csv_export(data)
        >>> "name" in csv_str
        True
    """
    try:
        output = StringIO()
        writer = None

        # Process tabular data if available
        if "rows" in data and isinstance(data["rows"], list) and len(data["rows"]) > 0:
            rows = data["rows"]

            if include_header and len(rows) > 0:
                # Extract headers from first row keys
                headers = list(rows[0].keys())
                writer = csv.DictWriter(output, fieldnames=headers, delimiter=delimiter)
                writer.writeheader()

            # Write data rows
            if writer:
                for row in rows:
                    writer.writerow(row)
            else:
                # Write without headers
                if len(rows) > 0:
                    headers = list(rows[0].keys())
                    writer = csv.DictWriter(output, fieldnames=headers, delimiter=delimiter)
                    for row in rows:
                        writer.writerow(row)

        else:
            # Create CSV from metrics and dimensions
            writer = csv.writer(output, delimiter=delimiter)

            # Write header if requested
            if include_header:
                writer.writerow(["Category", "Field", "Value"])

            # Write metrics
            if "metrics" in data:
                for metric_name, metric_value in data["metrics"].items():
                    if isinstance(metric_value, dict):
                        # Handle nested metrics
                        for k, v in metric_value.items():
                            writer.writerow(["Metric", f"{metric_name}.{k}", v])
                    else:
                        writer.writerow(["Metric", metric_name, metric_value])

            # Write dimensions
            if "dimensions" in data:
                for dim_name, dim_data in data["dimensions"].items():
                    if isinstance(dim_data, dict):
                        for k, v in dim_data.items():
                            writer.writerow(["Dimension", f"{dim_name}.{k}", v])
                    else:
                        writer.writerow(["Dimension", dim_name, dim_data])

        csv_content = output.getvalue()
        logger.info(f"CSV file generated ({len(csv_content)} characters)")
        return csv_content

    except Exception as e:
        logger.error(f"Failed to generate CSV file: {e}", exc_info=True)
        return None


@shared_task(
    name="tasks.export_tasks.export_to_excel",
    bind=True,
    max_retries=2,
    default_retry_delay=60,
)
def export_to_excel(
    self,
    data: Dict[str, Any],
    sheet_name: str = "Analytics Data",
    include_metadata: bool = True,
) -> Dict[str, Any]:
    """
    Export analytics data to Excel format.

    This Celery task handles Excel export from analytics data:
    1. Validate data structure
    2. Generate Excel workbook with data
    3. Optionally add metadata sheet
    4. Return Excel file as bytes

    Task Workflow:
    1. Validate data
    2. Generate Excel workbook
    3. Add styling and formatting
    4. Calculate file size and metadata
    5. Return Excel bytes with results

    Args:
        self: Celery task instance (bind=True)
        data: Analytics data dictionary containing:
            - metrics: Dictionary of metric name -> value
            - dimensions: Dictionary of dimension name -> data
            - summary: Text summary of the report
            - rows: Optional list of row data for tabular export
            - report_name: Optional report name
            - generated_at: Optional ISO timestamp
        sheet_name: Name of the main data sheet
        include_metadata: Whether to include metadata sheet

    Returns:
        Dictionary containing export results:
        - status: Task status (completed/failed)
        - excel_bytes: Excel file as bytes (if successful)
        - file_size_bytes: Size of Excel file in bytes
        - format: File format ("excel")
        - sheet_name: Name of the data sheet
        - processing_time_ms: Total processing time
        - error: Error message (if failed)

    Raises:
        SoftTimeLimitExceeded: If task exceeds time limit
        ValueError: If data is invalid
        Exception: For Excel generation errors

    Example:
        >>> from tasks.export_tasks import export_to_excel
        >>> data = {
        ...     "metrics": {"time_to_hire": 15.2},
        ...     "rows": [{"name": "John", "value": 100}]
        ... }
        >>> task = export_to_excel.delay(data)
        >>> result = task.get()
        >>> result['status']
        'completed'
    """
    start_time = time.time()
    total_steps = 3
    current_step = 0

    try:
        logger.info(f"Starting Excel export, sheet: {sheet_name}")

        # Step 1: Validate data
        current_step += 1
        progress = {
            "current": current_step,
            "total": total_steps,
            "percentage": int(current_step / total_steps * 100),
            "status": "validating_data",
            "message": "Validating export data...",
        }
        self.update_state(state="PROGRESS", meta=progress)
        logger.info(
            f"Task {self.request.id}: Step {current_step}/{total_steps} - "
            "Validating data"
        )

        if not data:
            raise ValueError("data cannot be empty")

        if not isinstance(data, dict):
            raise ValueError("data must be a dictionary")

        # Ensure generated_at exists
        if "generated_at" not in data:
            data["generated_at"] = datetime.utcnow().isoformat()

        logger.info("Export data validated successfully")

        # Step 2: Generate Excel file
        current_step += 1
        progress = {
            "current": current_step,
            "total": total_steps,
            "percentage": int(current_step / total_steps * 100),
            "status": "generating_excel",
            "message": "Generating Excel file...",
        }
        self.update_state(state="PROGRESS", meta=progress)
        logger.info(
            f"Task {self.request.id}: Step {current_step}/{total_steps} - "
            "Generating Excel"
        )

        excel_bytes = generate_excel_export(
            data=data,
            sheet_name=sheet_name,
            include_metadata=include_metadata,
        )

        if excel_bytes is None:
            raise Exception("Excel generation failed")

        logger.info(f"Excel file generated ({len(excel_bytes)} bytes)")

        # Step 3: Prepare results
        current_step += 1
        progress = {
            "current": current_step,
            "total": total_steps,
            "percentage": int(current_step / total_steps * 100),
            "status": "finalizing",
            "message": "Finalizing Excel export...",
        }
        self.update_state(state="PROGRESS", meta=progress)
        logger.info(
            f"Task {self.request.id}: Step {current_step}/{total_steps} - "
            "Finalizing"
        )

        processing_time_ms = round((time.time() - start_time) * 1000, 2)

        result = {
            "status": "completed",
            "excel_bytes": excel_bytes,
            "file_size_bytes": len(excel_bytes),
            "format": "excel",
            "sheet_name": sheet_name,
            "include_metadata": include_metadata,
            "processing_time_ms": processing_time_ms,
            "task_id": self.request.id,
        }

        logger.info(
            f"Excel export completed: size: {len(excel_bytes)} bytes, "
            f"time: {processing_time_ms}ms"
        )

        return result

    except SoftTimeLimitExceeded:
        logger.error(f"Task {self.request.id} exceeded time limit")
        return {
            "status": "failed",
            "error": "Excel export exceeded maximum time limit",
            "format": "excel",
            "processing_time_ms": round((time.time() - start_time) * 1000, 2),
        }

    except ValueError as e:
        logger.error(f"Validation error in Excel export: {e}")
        return {
            "status": "failed",
            "error": f"Validation error: {str(e)}",
            "format": "excel",
            "processing_time_ms": round((time.time() - start_time) * 1000, 2),
        }

    except Exception as e:
        logger.error(f"Error in Excel export: {e}", exc_info=True)
        return {
            "status": "failed",
            "error": str(e),
            "format": "excel",
            "processing_time_ms": round((time.time() - start_time) * 1000, 2),
        }


@shared_task(
    name="tasks.export_tasks.export_to_csv",
    bind=True,
    max_retries=2,
    default_retry_delay=60,
)
def export_to_csv(
    self,
    data: Dict[str, Any],
    delimiter: str = ",",
    include_header: bool = True,
) -> Dict[str, Any]:
    """
    Export analytics data to CSV format.

    This Celery task handles CSV export from analytics data:
    1. Validate data structure
    2. Generate CSV content
    3. Return CSV as string

    Task Workflow:
    1. Validate data
    2. Generate CSV content
    3. Calculate file size and metadata
    4. Return CSV content with results

    Args:
        self: Celery task instance (bind=True)
        data: Analytics data dictionary containing:
            - metrics: Dictionary of metric name -> value
            - dimensions: Dictionary of dimension name -> data
            - rows: Optional list of row data for tabular export
        delimiter: CSV delimiter (default: comma)
        include_header: Whether to include header row

    Returns:
        Dictionary containing export results:
        - status: Task status (completed/failed)
        - csv_content: CSV content as string (if successful)
        - file_size_bytes: Size of CSV content in bytes
        - format: File format ("csv")
        - delimiter: Delimiter used
        - processing_time_ms: Total processing time
        - error: Error message (if failed)

    Raises:
        SoftTimeLimitExceeded: If task exceeds time limit
        ValueError: If data is invalid
        Exception: For CSV generation errors

    Example:
        >>> from tasks.export_tasks import export_to_csv
        >>> data = {
        ...     "rows": [{"name": "John", "value": 100}]
        ... }
        >>> task = export_to_csv.delay(data)
        >>> result = task.get()
        >>> result['status']
        'completed'
    """
    start_time = time.time()
    total_steps = 3
    current_step = 0

    try:
        logger.info(f"Starting CSV export, delimiter: {delimiter}")

        # Step 1: Validate data
        current_step += 1
        progress = {
            "current": current_step,
            "total": total_steps,
            "percentage": int(current_step / total_steps * 100),
            "status": "validating_data",
            "message": "Validating export data...",
        }
        self.update_state(state="PROGRESS", meta=progress)
        logger.info(
            f"Task {self.request.id}: Step {current_step}/{total_steps} - "
            "Validating data"
        )

        if not data:
            raise ValueError("data cannot be empty")

        if not isinstance(data, dict):
            raise ValueError("data must be a dictionary")

        logger.info("Export data validated successfully")

        # Step 2: Generate CSV content
        current_step += 1
        progress = {
            "current": current_step,
            "total": total_steps,
            "percentage": int(current_step / total_steps * 100),
            "status": "generating_csv",
            "message": "Generating CSV content...",
        }
        self.update_state(state="PROGRESS", meta=progress)
        logger.info(
            f"Task {self.request.id}: Step {current_step}/{total_steps} - "
            "Generating CSV"
        )

        csv_content = generate_csv_export(
            data=data,
            delimiter=delimiter,
            include_header=include_header,
        )

        if csv_content is None:
            raise Exception("CSV generation failed")

        logger.info(f"CSV content generated ({len(csv_content)} characters)")

        # Step 3: Prepare results
        current_step += 1
        progress = {
            "current": current_step,
            "total": total_steps,
            "percentage": int(current_step / total_steps * 100),
            "status": "finalizing",
            "message": "Finalizing CSV export...",
        }
        self.update_state(state="PROGRESS", meta=progress)
        logger.info(
            f"Task {self.request.id}: Step {current_step}/{total_steps} - "
            "Finalizing"
        )

        processing_time_ms = round((time.time() - start_time) * 1000, 2)

        result = {
            "status": "completed",
            "csv_content": csv_content,
            "file_size_bytes": len(csv_content.encode('utf-8')),
            "format": "csv",
            "delimiter": delimiter,
            "include_header": include_header,
            "processing_time_ms": processing_time_ms,
            "task_id": self.request.id,
        }

        logger.info(
            f"CSV export completed: size: {len(csv_content)} characters, "
            f"time: {processing_time_ms}ms"
        )

        return result

    except SoftTimeLimitExceeded:
        logger.error(f"Task {self.request.id} exceeded time limit")
        return {
            "status": "failed",
            "error": "CSV export exceeded maximum time limit",
            "format": "csv",
            "processing_time_ms": round((time.time() - start_time) * 1000, 2),
        }

    except ValueError as e:
        logger.error(f"Validation error in CSV export: {e}")
        return {
            "status": "failed",
            "error": f"Validation error: {str(e)}",
            "format": "csv",
            "processing_time_ms": round((time.time() - start_time) * 1000, 2),
        }

    except Exception as e:
        logger.error(f"Error in CSV export: {e}", exc_info=True)
        return {
            "status": "failed",
            "error": str(e),
            "format": "csv",
            "processing_time_ms": round((time.time() - start_time) * 1000, 2),
        }


@shared_task(
    name="tasks.export_tasks.batch_export_reports",
    bind=True,
    max_retries=2,
)
def batch_export_reports(
    self,
    exports: List[Dict[str, Any]],
) -> Dict[str, Any]:
    """
    Export multiple reports in batch.

    This task handles batch export for multiple reports in various formats,
    useful for scheduled report delivery or bulk export operations.

    Args:
        self: Celery task instance (bind=True)
        exports: List of export configurations, each containing:
            - data: Analytics data dictionary
            - format: Export format ("excel" or "csv")
            - options: Optional format-specific options

    Returns:
        Dictionary containing batch results:
        - total_exports: Total number of exports requested
        - successful: Number of successfully generated exports
        - failed: Number of failed exports
        - results: List of individual results
        - processing_time_ms: Total processing time

    Example:
        >>> from tasks.export_tasks import batch_export_reports
        >>> exports = [
        ...     {"data": {...}, "format": "excel"},
        ...     {"data": {...}, "format": "csv"}
        ... ]
        >>> task = batch_export_reports.delay(exports)
        >>> result = task.get()
        >>> result['successful']
        2
    """
    start_time = time.time()

    try:
        logger.info(f"Starting batch export for {len(exports)} reports")

        total_exports = len(exports)
        successful = 0
        failed = 0
        results = []

        for idx, export_config in enumerate(exports):
            try:
                # Update progress
                progress = {
                    "current": idx + 1,
                    "total": total_exports,
                    "percentage": int((idx + 1) / total_exports * 100),
                    "status": "processing",
                    "message": f"Processing export {idx + 1}/{total_exports}...",
                }
                self.update_state(state="PROGRESS", meta=progress)

                # Extract configuration
                data = export_config.get("data", {})
                export_format = export_config.get("format", "csv").lower()
                options = export_config.get("options", {})

                # Generate export based on format
                if export_format == "excel":
                    result = export_to_excel(
                        data=data,
                        sheet_name=options.get("sheet_name", "Analytics Data"),
                        include_metadata=options.get("include_metadata", True),
                    )
                elif export_format == "csv":
                    result = export_to_csv(
                        data=data,
                        delimiter=options.get("delimiter", ","),
                        include_header=options.get("include_header", True),
                    )
                else:
                    raise ValueError(f"Unsupported export format: {export_format}")

                if result.get("status") == "completed":
                    successful += 1
                    logger.info(f"Successfully generated {export_format} export {idx + 1}")
                else:
                    failed += 1
                    logger.warning(f"Failed to generate {export_format} export {idx + 1}")

                results.append({
                    "export_index": idx + 1,
                    "format": export_format,
                    "status": result.get("status"),
                    "file_size_bytes": result.get("file_size_bytes"),
                    "error": result.get("error"),
                })

            except Exception as e:
                failed += 1
                logger.error(f"Error processing export {idx + 1}: {e}")
                results.append({
                    "export_index": idx + 1,
                    "format": export_config.get("format", "unknown"),
                    "status": "failed",
                    "error": str(e),
                })

        processing_time_ms = round((time.time() - start_time) * 1000, 2)

        summary = {
            "total_exports": total_exports,
            "successful": successful,
            "failed": failed,
            "results": results,
            "processing_time_ms": processing_time_ms,
            "status": "completed",
        }

        logger.info(
            f"Batch export completed: {successful} successful, "
            f"{failed} failed, time: {processing_time_ms}ms"
        )

        return summary

    except Exception as e:
        logger.error(f"Error in batch export: {e}", exc_info=True)
        return {
            "total_exports": len(exports),
            "successful": 0,
            "failed": len(exports),
            "results": [],
            "processing_time_ms": round((time.time() - start_time) * 1000, 2),
            "status": "failed",
            "error": str(e),
        }
